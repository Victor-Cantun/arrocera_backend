from cgi import print_arguments
from ipaddress import summarize_address_range

# from msilib.schema import ListView
from django.views.generic import ListView, View, TemplateView
from urllib import request
from django.shortcuts import render, redirect
from .models import (
    BankAccountsCustomer,
    BankAccountsEmployee,
    BankAccountsProvider,
    BillOfLading,
    BilledIncome,
    BillsPaidPlugins,
    ChargerSalary,
    Company,
    ConceptPayment,
    Conciliation,
    CreditNote,
    Customer,
    Department,
    DepositControl,
    DoubleDays,
    Driver,
    DriverSalary,
    ExtraHours,
    FileProducer,
    FileQuotation,
    FileUnit,
    FuelDump,
    FuelType,
    Fueling,
    Gasoline,
    Income,
    IncomeCreditNotes,
    IncomeInvoices,
    InitialCash,
    LandRent,
    Loans,
    LoansChargers,
    LoansDrivers,
    Location,
    MainPresentation,
    MainProduct,
    Output,
    OutputProduct,
    Outputreview,
    PaidPlugins,
    Parcel,
    PaymentOrderProducer,
    PaymentProducer,
    PaymentSchedule,
    PaymentsChargers,
    PaymentsDrivers,
    Payroll,
    PettyCash,
    Plate,
    Presentation,
    Producer,
    Product,
    ProductBI,
    ProductCN,
    ProductPO,
    ProductQuotation,
    ProductRequisition,
    ProductShopping,
    ProductW,
    Props,
    Provider,
    Category,
    Employee,
    PurchaseOrder,
    Quotation,
    Requisition,
    Rowoutputreview,
    Sale,
    SegalmexParcel,
    SegalmexReception,
    Shopping,
    Society,
    SocietyProducers,
    User,
    Bank,
    BankAccount,
    UploadImage,
    Unit,
    Variety,
    VehicleType,
    Warehouse,
    Ticketreview,
    Rowticketreview,
    Binnacle,
)

# RestFramework
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import generics
from rest_framework import viewsets
from num2words import num2words
from .serializers import (
    BankAccountsCustomerSerializer,
    BankAccountsEmployeeSerializer,
    BankAccountsProviderSerializer,
    BillOfLadingSerializer,
    BilledIncomeSerializer,
    BinnacleSerializer,
    ChargerSalarySerializer,
    CompanySerializer,
    ConciliationSerializer,
    CreditNoteListSerializer,
    CustomerSerializer,
    DepositControlSerializer,
    DieselSerializer,
    DoubleDaysSerializer,
    DriverSalarySerializer,
    DriverSerializer,
    ExtraHoursSerializer,
    FilePaidPluginsSerializer,
    FileProducerSerializer,
    FileQuotationSerializer,
    FileUnitSerializer,
    FuelDumpListSerializer,
    FuelDumpSerializer,
    FuelTypeSerializer,
    FuelingListSerializer,
    FuelingListSerializer,
    FuelingSerializer,
    IncomeSerializer,
    InitialCashSerializer,
    LandRentSerializer,
    ListAccountStatusSerializer,
    ListBankAccountsCustomerSerializer,
    ListBankAccountsEmployeeSerializer,
    ListBankAccountsProviderSerializer,
    ListBillOfLadingSerializer,
    ListBilledIncomeSerializer,
    ListCalculatePayrollSerializer,
    ListChargerSalarySerializer,
    ListConciliationSerializer,
    ListDepositControlSerializer,
    ListDoubleDaysSerializer,
    ListDriverSalarySerializer,
    ListExtraHoursSerializer,
    ListIncomeSerializer,
    ListLandRentSerializer,
    ListLoansChargersSerializer,
    ListLoansDriversSerializer,
    ListLoansSerializer,
    ListPaidPluginsSerializer,
    ListPaymentOrderProducerSerializer,
    ListPaymentProducerSerializer,
    ListPaymentScheduleSerializer,
    ListPaymentsChargersSerializer,
    ListPaymentsDriversSerializer,
    ListPayrollSerializer,
    ListPettyCashSerializer,
    ListPropsSerializer,
    ListPurchaseOrdersSerializers,
    ListQuotationSerializer,
    ListSalesSerializer,
    ListSegalmexParcelSerializer,
    ListSegalmexReceptionSerializer,
    ListTotalDepositControlSerializer,
    LoansChargersSerializer,
    LoansDriversSerializer,
    LoansSerializer,
    LocationSerializer,
    MainPresentationSerializer,
    MainProductSerializer,
    OutputListSerializer,
    OutputreviewListSerializer,
    PaidPluginsSerializer,
    ParcelListSerializer,
    ParcelSerializer,
    PaymentOrderProducerSerializer,
    PaymentProducerSerializer,
    PaymentsChargersSerializer,
    PaymentsDriversSerializer,
    PaymentsSerializer,
    PayrollSerializer,
    PettyCashSerializer,
    PlateSerializer,
    PresentationSerializer,
    ProductSerializer,
    PropsSerializer,
    PurchaseOrderListSerializer,
    RequisitionListSerializer,
    RequisitionSerializer,
    SaleSerializer,
    SegalmexParcelSerializer,
    SegalmexReceptionSerializer,
    ShoppingListSerializer,
    SocietyListSerializer,
    SocietySerializer,
    UnitListSerializer,
    UserSerializer,
    DepartmentSerializer,
    BankSerializer,
    BankAccountSerializer,
    EmployeeSerializer,
    EmployeeListSerializer,
    ProducerSerializer,
    ProviderSerializer,
    CategorySerializer,
    BankAccountListSerializer,
    UploadImageSerializer,
    UnitSerializer,
    VarietySerializer,
    VehicleTypeSerializer,
    WarehouseListSerializer,
    WarehouseSerializer,
    TicketreviewListSerializer,
)

# FORMS
from .forms import EmployeeForm
from django.http import HttpResponse, JsonResponse
from .utils import render_to_pdf
import openpyxl

# from openpyxl.styles.borders import Border, Side
# from openpyxl.styles import Style
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from werkzeug.security import generate_password_hash, check_password_hash
from django.db import models
from django.db.models import (
    Sum,
    F,
    Value,
    CharField,
    Avg,
    Subquery,
    OuterRef,
    Exists,
    Min,
)
from django.db.models.functions import Concat, Coalesce
from django.db.models import Count

# Create your views here.
from io import BytesIO
from xhtml2pdf import pisa
import json
from django.core.serializers import serialize


def home(request):
    return render(request, "home.html")


# *? USUARIOS
@api_view(["POST"])
def RegisterUser(request):
    data = request.data
    password = data["password"]
    clave = generate_password_hash(password)
    new_user = User.objects.create(
        name=data["name"],
        lastname=data["lastname"],
        cellphone=data["cellphone"],
        email=data["email"],
        password=clave,
    )
    new_user.save()
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["POST"])
def LoginUser(request):
    data = request.data
    clave = data["password"]
    users = User.objects.filter(email=data["email"]).first()
    if not users:
        return Response(
            {
                "message": "No existe coincidencias con su correo electrónico!",
                "status": 400,
            }
        )
    if check_password_hash(users.password, clave):
        serializer = UserSerializer(users, many=False)
        return Response(serializer.data)
    return Response({"message": "No existe coincidencias!", "status": 400})


@api_view(["POST"])
def CreateUser(request):
    data = request.data
    password = data["password"]
    clave = generate_password_hash(password)
    new_user = User.objects.create(
        name=data["name"],
        lastname=data["lastname"],
        cellphone=data["cellphone"],
        email=data["email"],
        password=clave,
        users=data["users"],
        companies=data["companies"],
        departments=data["departments"],
        categories=data["categories"],
        societies=data["societies"],
        parcels=data["parcels"],
        producers=data["producers"],
        reception=data["reception"],
        binnacle=data["binnacle"],
        units=data["units"],
        fuels=data["fuels"],
        requisitions=data["requisitions"],
        shoppings=data["shoppings"],
        providers=data["providers"],
        quotes=data["quotes"],
        purchase_orders=data["purchase_orders"],
        products=data["products"],
        outputs=data["outputs"],
        CreditNotes=data["CreditNotes"],
        BillOfLading=data["BillOfLading"],
        PaymentSchedule=data["PaymentSchedule"],
        PettyCash=data["PettyCash"],
        drivers=data["drivers"],
        chargers=data["chargers"],
        LandRent=data["LandRent"],
        PaymentOrderProducer=data["PaymentOrderProducer"],
        payments=data["payments"],
        banks=data["banks"],
        customers=data["customers"],
        BilledIncome=data["BilledIncome"],
        PaidPlugins=data["PaidPlugins"],
        DepositControl=data["DepositControl"],
        Income=data["Income"],
        Conciliation=data["Conciliation"],
        employees=data["employees"],
        payroll=data["payroll"],
        ExtraHours=data["ExtraHours"],
        DoubleDays=data["DoubleDays"],
        props=data["props"],
        loans=data["loans"],
        watch=data["watch"],
        write=data["write"],
        edit=data["edit"],
        delete=data["delete"],
        export=data["export"],
        status=data["status"],
        role=data["role"],
    )
    new_user.save()
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["GET"])
def ListUser(request):
    users = User.objects.all()
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def DetailUser(request, pk):
    user = User.objects.get(id=pk)
    serializer = UserSerializer(user, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateUser(request, pk):
    user = User.objects.get(id=pk)
    serializer = UserSerializer(instance=user, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdatePassword(request):
    if (
        ("email" in request.data)
        and ("old_password" in request.data)
        and ("password" in request.data)
    ):
        email = request.data["email"]
        old_password = request.data["old_password"]
        new_password = generate_password_hash(request.data["password"])
        user = User.objects.filter(email=email).first()
        if check_password_hash(user.password, old_password):
            User.objects.filter(email=email).update(password=new_password)
            return Response(
                {
                    "message": "Coontraseña actualizada satisfactoriamente!",
                    "status": 200,
                }
            )
        return Response({"message": "Contraseña incorrecta!", "status": 400})


@api_view(["POST"])
def UpdatePhotoUser(request, pk):
    data = request.data
    employee = User.objects.get(id=pk)
    employee.photo = data["imagen"]
    employee.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["DELETE"])
def DeleteUser(request, pk):
    row = User.objects.filter(id=pk)
    row.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


# *? COMPAÑIAS
@api_view(["GET"])
def ListCompany(request):
    company = Company.objects.all()
    serializer = CompanySerializer(company, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateCompany(request):
    serializer = CompanySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteCompany(request, pk):
    company = Company.objects.get(id=pk)
    company.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


@api_view(["GET"])
def DetailCompany(request, pk):
    company = Company.objects.get(id=pk)
    serializer = CompanySerializer(company, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateCompany(request, pk):
    company = Company.objects.get(id=pk)
    serializer = DepartmentSerializer(instance=company, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? DEPARTAMENTOS


@api_view(["GET"])
def DepartmentList(request):
    department = Department.objects.all().order_by("id")
    serializer = DepartmentSerializer(department, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def DepartmentDetail(request, pk):
    department = Department.objects.get(id=pk)
    serializer = DepartmentSerializer(department, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateDepartment(request):
    serializer = DepartmentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateDepartment(request, pk):
    department = Department.objects.get(id=pk)
    serializer = DepartmentSerializer(instance=department, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteDepartment(request, pk):
    department = Department.objects.get(id=pk)
    department.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


class ReportDepartmentsPDF(View):
    def get(self, request, *args, **kwargs):
        departments = Department.objects.all()
        data = {"departments": departments}
        pdf = render_to_pdf("paginas/ReportDepartments.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportDepartmentsXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        departments = Department.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Categorias"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Departamento"

        cont = 7
        pos = 1
        for department in departments:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = department.name

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_departamentos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? CATEGORIAS


@api_view(["GET"])
def CategoryList(request):
    category = Category.objects.all().order_by("id")
    serializer = CategorySerializer(category, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def CategoryDetail(request, pk):
    category = Category.objects.get(id=pk)
    serializer = CategorySerializer(category, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateCategory(request):
    serializer = CategorySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateCategory(request, pk):
    category = Category.objects.get(id=pk)
    serializer = CategorySerializer(instance=category, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteCategory(request, pk):
    category = Category.objects.get(id=pk)
    category.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


class ReportCategoriesPDF(View):
    def get(self, request, *args, **kwargs):
        categories = Category.objects.all()
        data = {"categories": categories}
        pdf = render_to_pdf("paginas/ReportCategories.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportCategoriesXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        categories = Category.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Categorias"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Categoria"

        cont = 7
        pos = 1
        for category in categories:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = category.name

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_categorias.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? BANCOS


@api_view(["GET"])
def BankList(request):
    banks = Bank.objects.all().order_by("id")
    serializer = BankSerializer(banks, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def BankDetail(request, pk):
    bank = Bank.objects.get(id=pk)
    serializer = BankSerializer(bank, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def BankCreate(request):
    serializer = BankSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def BankUpdate(request, pk):
    bank = Bank.objects.get(id=pk)
    serializer = BankSerializer(instance=bank, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def BankDelete(request, pk):
    bank = Bank.objects.get(id=pk)
    bank.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


class ReportBanksPDF(View):
    def get(self, request, *args, **kwargs):
        banks = Bank.objects.all()
        data = {"banks": banks}
        pdf = render_to_pdf("paginas/ReportBanks.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportBanksXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        banks = Bank.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Bancos"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Banco"

        cont = 7
        pos = 1
        for bank in banks:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = bank.name

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_bancos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? CUENTAS DE BANCOS


@api_view(["GET"])
def BankAccountList(request):
    banksaccount = BankAccount.objects.all().order_by("id")
    serializer = BankAccountListSerializer(banksaccount, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def BankAccountDetail(request, pk):
    bankaccount = BankAccount.objects.get(id=pk)
    serializer = BankAccountSerializer(bankaccount, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def BankAccountCreate(request):
    serializer = BankAccountSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def BankAccountUpdate(request, pk):
    bankaccount = BankAccount.objects.get(id=pk)
    serializer = BankAccountSerializer(instance=bankaccount, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def BankAccountDelete(request, pk):
    bankaccount = BankAccount.objects.get(id=pk)
    bankaccount.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


class ReportBankAccountsPDF(View):
    def get(self, request, *args, **kwargs):
        bankaccounts = BankAccount.objects.all()
        data = {"bankaccounts": bankaccounts}
        pdf = render_to_pdf("paginas/ReportBankAccounts.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportBankAccountsXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        bankaccounts = BankAccount.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Bancos"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Banco"
        ws["E6"] = "No. Cuenta"
        ws["F6"] = "Clave interbancaria"

        cont = 7
        pos = 1
        for bankaccount in bankaccounts:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = bankaccount.bank.name
            ws.cell(row=cont, column=5).value = bankaccount.number
            ws.cell(row=cont, column=6).value = bankaccount.interbank_key

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_cuentas_de_banco.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? PRODUCTORES
@api_view(["POST"])
def Producers(request):
    if "user" in request.data:
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Producer.objects.all()

        # *? Empleado
        if user != 0:
            rows = Producer.objects.filter(user_id=user)

        serializer = ProducerSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ProducerList(request):
    producer = Producer.objects.all()
    serializer = ProducerSerializer(producer, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def ProducerDetail(request, pk):
    producer = Producer.objects.get(id=pk)
    serializer = ProducerSerializer(producer, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateProducer(request):
    serializer = ProducerSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateProducer(request, pk):
    producer = Producer.objects.get(id=pk)
    serializer = ProducerSerializer(instance=producer, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteProducer(request, pk):
    producer = Producer.objects.get(id=pk)
    producer.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? PDF PRODUCTORES
# class ReportProducersPDF(View):
@api_view(["GET"])
def ReportProducersPDF(request, user):
    user = int(user)
    # return Response({"department->":department,"user->":user})
    # *? administrador
    if user == 0:
        rows = Producer.objects.all()

    # *? Empleado
    if user != 0:
        rows = Producer.objects.filter(user_id=user)

    # producers = Producer.objects.all()

    data = {"producers": rows}
    pdf = render_to_pdf("paginas/ReportProducer.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


# *? XLSX PRODUCTORES
class ReportProducersXLSX(TemplateView):
    def get(self, request, user):
        user = int(user)
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Producer.objects.all()

        # *? Empleado
        if user != 0:
            rows = Producer.objects.filter(user_id=user)

        producers = rows
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Productores"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Nombre"
        ws["E6"] = "RFC"
        ws["F6"] = "Teléfono"
        ws["G6"] = "Correo electrónico"
        ws["H6"] = "Código Postal"
        ws["I6"] = "Pais"
        ws["J6"] = "Estado"
        ws["K6"] = "Ciudad"
        ws["L6"] = "Localidad"
        ws["M6"] = "Colonia"
        ws["N6"] = "Calle"
        ws["O6"] = "No. Interior"
        ws["P6"] = "No. Exterior"
        ws["Q6"] = "Representante legal"

        cont = 7

        for producer in producers:
            ws.cell(row=cont, column=3).value = producer.id
            ws.cell(row=cont, column=4).value = producer.name
            ws.cell(row=cont, column=5).value = producer.rfc
            ws.cell(row=cont, column=6).value = producer.phone
            ws.cell(row=cont, column=7).value = producer.email
            ws.cell(row=cont, column=8).value = producer.postal_code
            ws.cell(row=cont, column=9).value = producer.country
            ws.cell(row=cont, column=10).value = producer.state
            ws.cell(row=cont, column=11).value = producer.city
            ws.cell(row=cont, column=12).value = producer.location
            ws.cell(row=cont, column=13).value = producer.suburb
            ws.cell(row=cont, column=14).value = producer.street
            ws.cell(row=cont, column=15).value = producer.int_no
            ws.cell(row=cont, column=16).value = producer.ext_no
            ws.cell(row=cont, column=17).value = producer.representative

            cont += 1

        nombre_archivo = "Reporte_Productores.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? PROVEEDORES
@api_view(["POST"])
def Providers(request):
    if "user" in request.data:
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Provider.objects.all()

        # *? Empleado
        if user != 0:
            rows = Provider.objects.filter(user_id=user)

        serializer = ProviderSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ProviderList(request):
    provider = Provider.objects.all()
    serializer = ProviderSerializer(provider, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def ProviderListForUser(request, pk):
    provider = Provider.objects.filter(user_id=pk)
    serializer = ProviderSerializer(provider, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def ProviderDetail(request, pk):
    provider = Provider.objects.get(id=pk)
    serializer = ProviderSerializer(provider, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateProvider(request):
    serializer = ProviderSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateProvider(request, pk):
    provider = Provider.objects.get(id=pk)
    serializer = ProviderSerializer(instance=provider, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteProvider(request, pk):
    provider = Provider.objects.get(id=pk)
    provider.delete()
    return Response("Registro eliminado satisfactoriamente!")


# class ReportProvidersPDF(View):
@api_view(["GET"])
def ReportProvidersPDF(request, user):
    user = int(user)
    # *? administrador
    if user == 0:
        rows = Provider.objects.all()
    # *? Empleado
    if user != 0:
        rows = Provider.objects.filter(user_id=user)
    data = {"providers": rows}
    pdf = render_to_pdf("paginas/ReportProviders.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportProvidersXLSX(TemplateView):
    def get(self, request, user):
        user = int(user)
        # *? administrador
        if user == 0:
            rows = Provider.objects.all()
        # *? Empleado
        if user != 0:
            rows = Provider.objects.filter(user_id=user)
        # providers = Provider.objects.all()
        providers = rows

        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Proveedores"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Nombre"
        ws["E6"] = "RFC"
        ws["F6"] = "Teléfono"
        ws["G6"] = "Correo electrónico"
        ws["H6"] = "Representante"
        ws["I6"] = "Codigo Postal"
        ws["J6"] = "Pais"
        ws["K6"] = "Estado"
        ws["L6"] = "Ciudad"
        ws["M6"] = "Localidad"
        ws["N6"] = "Colonia"
        ws["O6"] = "Calle"
        ws["P6"] = "No. Interior"
        ws["Q6"] = "No. Exterior"

        cont = 7
        pos = 1

        for provider in providers:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = provider.name
            ws.cell(row=cont, column=5).value = provider.rfc
            ws.cell(row=cont, column=6).value = provider.phone
            ws.cell(row=cont, column=7).value = provider.email
            ws.cell(row=cont, column=8).value = provider.representative
            ws.cell(row=cont, column=9).value = provider.postal_code
            ws.cell(row=cont, column=10).value = provider.country
            ws.cell(row=cont, column=11).value = provider.state
            ws.cell(row=cont, column=12).value = provider.city
            ws.cell(row=cont, column=13).value = provider.location
            ws.cell(row=cont, column=14).value = provider.suburb
            ws.cell(row=cont, column=15).value = provider.street
            ws.cell(row=cont, column=16).value = provider.int_no
            ws.cell(row=cont, column=17).value = provider.ext_no
            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Proveedores.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? EMPLEADOS
@api_view(["POST"])
def Employees(request):
    if ("department" in request.data) and ("user" in request.data):
        department = int(request.data["department"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (department == 0) and (user == 0):
            employee = (
                Employee.objects.all()
                .annotate(
                    nombre_completo=Concat(
                        "surname",
                        Value(" "),
                        "second_surname",
                        Value(" "),
                        "name",
                        output_field=models.CharField(),
                    )
                )
                .order_by("surname")
            )

        if (department != 0) and (user == 0):
            employee = (
                Employee.objects.filter(department_id=department)
                .annotate(
                    nombre_completo=Concat(
                        "surname",
                        Value(" "),
                        "second_surname",
                        Value(" "),
                        "name",
                        output_field=models.CharField(),
                    )
                )
                .order_by("surname")
            )

        # *? Empleado
        if (department == 0) and (user != 0):
            employee = (
                Employee.objects.filter(user_id=user)
                .annotate(
                    nombre_completo=Concat(
                        "surname",
                        Value(" "),
                        "second_surname",
                        Value(" "),
                        "name",
                        output_field=models.CharField(),
                    )
                )
                .order_by("surname")
            )

        if (department != 0) and (user != 0):
            employee = (
                Employee.objects.filter(user_id=user)
                .filter(department_id=department)
                .annotate(
                    nombre_completo=Concat(
                        "surname",
                        Value(" "),
                        "second_surname",
                        Value(" "),
                        "name",
                        output_field=models.CharField(),
                    )
                )
                .order_by("surname")
            )

        serializer = EmployeeListSerializer(employee, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListEmployee(request):
    employee = Employee.objects.all()
    serializer = EmployeeListSerializer(employee, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def EmployeeListForUser(request, pk):
    employee = (
        Employee.objects.filter(user_id=pk)
        .annotate(
            nombre_completo=Concat(
                "surname",
                Value(" "),
                "second_surname",
                Value(" "),
                "name",
                output_field=models.CharField(),
            )
        )
        .order_by("surname")
    )
    serializer = EmployeeListSerializer(employee, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def EmployeeDetail(request, pk):
    employee = Employee.objects.get(id=pk)
    serializer = EmployeeListSerializer(employee, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateEmployee(request):
    if request.method == "POST":
        serializer = EmployeeSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"message": "Registro agregado satisfactoriamente!", "status": 200}
            )
        return Response(
            {
                "message": "No se realizó el Registro!",
                "errors": serializer.errors,
                "status": 400,
            }
        )


@api_view(["POST"])
def UpdateEmployee(request, pk):
    employee = Employee.objects.get(id=pk)
    serializer = EmployeeSerializer(instance=employee, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdatePhotoEmployee(request, pk):
    data = request.data
    employee = Employee.objects.get(id=pk)
    employee.imagen = data["imagen"]
    employee.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["DELETE"])
def DeleteEmployee(request, pk):
    employee = Employee.objects.get(id=pk)
    employee.delete()
    return Response("Registro eliminado satisfactoriamente!")


def CrearEmpleado(request):
    formulario = EmployeeForm(request.POST or None, request.FILES or None)
    if formulario.is_valid():
        formulario.save()
        return redirect("home")
    return render(request, "paginas/CreateEmployee.html", {"formulario": formulario})


class EmpleadoList(generics.ListCreateAPIView):
    queryset = Employee.objects.all()
    serializer_class = EmployeeSerializer


class ReportEmployee(ListView):
    model = Employee
    template_name = "paginas/ReportEmployee.html"
    context_object_name = "employees"


# class ReportEmployeesPDF(View):
@api_view(["GET"])
def ReportEmployeesPDF(request, department, user):
    department = int(department)
    user = int(user)
    # *? administrador
    if (department == 0) and (user == 0):
        rows = Employee.objects.all()
    if (department != 0) and (user == 0):
        rows = Employee.objects.filter(department_id=department)
    # *? Empleado
    if (department == 0) and (user != 0):
        rows = Employee.objects.filter(user_id=user)
    if (department != 0) and (user != 0):
        rows = Employee.objects.filter(user_id=user).filter(department_id=department)
    employees = rows
    data = {"employees": employees}
    pdf = render_to_pdf("paginas/ReportEmployees.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportEmployeesXLSX(TemplateView):
    def get(self, request, department, user):
        department = int(department)
        user = int(user)
        # *? administrador
        if (department == 0) and (user == 0):
            rows = Employee.objects.all()
        if (department != 0) and (user == 0):
            rows = Employee.objects.filter(department_id=department)
        # *? Empleado
        if (department == 0) and (user != 0):
            rows = Employee.objects.filter(user_id=user)
        if (department != 0) and (user != 0):
            rows = Employee.objects.filter(user_id=user).filter(
                department_id=department
            )
        employees = rows

        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Empleados"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Nombre"
        ws["E6"] = "Apellido paterno"
        ws["F6"] = "Appllido materno"
        ws["G6"] = "Departamento"
        ws["H6"] = "Categoria"
        ws["I6"] = "Correo electrónico"
        ws["J6"] = "Celular"

        cont = 7
        pos = 1

        for empleado in employees:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = empleado.name
            ws.cell(row=cont, column=5).value = empleado.surname
            ws.cell(row=cont, column=6).value = empleado.second_surname
            ws.cell(row=cont, column=7).value = (
                empleado.department.name if empleado.department != None else ""
            )
            ws.cell(row=cont, column=8).value = (
                empleado.category.name if empleado.category != None else ""
            )
            ws.cell(row=cont, column=9).value = empleado.personal_email
            ws.cell(row=cont, column=10).value = empleado.cell_phone
            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Empleados.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? UNIDADES
@api_view(["POST"])
def Units(request):
    if "user" in request.data:
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Unit.objects.all()

        # *? Empleado
        if user != 0:
            rows = Unit.objects.filter(user_id=user)

        serializer = UnitListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def VehicleTypeList(request):
    vehicles = VehicleType.objects.all()
    serializer = VehicleTypeSerializer(vehicles, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateVehicleType(request):
    serializer = VehicleTypeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def UnitList(request):
    units = Unit.objects.all()
    serializer = UnitListSerializer(units, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def UnitListForUser(request, pk):
    units = Unit.objects.filter(user_id=pk)
    serializer = UnitListSerializer(units, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def UnitDetail(request, pk):
    units = Unit.objects.get(id=pk)
    serializer = UnitListSerializer(units, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateUnit(request):
    serializer = UnitSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateUnit(request, pk):
    unit = Unit.objects.get(id=pk)
    serializer = UnitSerializer(instance=unit, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdatePhotoUnit(request, pk):
    data = request.data
    unit = Unit.objects.get(id=pk)
    unit.imagen = data["imagen"]
    unit.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )
    # return Response({"message":"No se realizó el Registro!","status":400})


@api_view(["DELETE"])
def DeleteUnit(request, pk):
    unit = Unit.objects.get(id=pk)
    unit.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["POST"])
def UploadFileUnit(request):
    serializer = FileUnitSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def DetailFileUnit(request, pk):
    documents = FileUnit.objects.filter(unit_id=pk)
    serializer = FileUnitSerializer(documents, many=True)
    return Response(serializer.data)


# class ReportUnitsPDF(View):
def ReportUnitsPDF(request, user):
    user = int(user)
    # *? administrador
    if user == 0:
        rows = Unit.objects.all()
    # *? Empleado
    if user != 0:
        rows = Unit.objects.filter(user_id=user)
    units = rows
    data = {"units": units}
    pdf = render_to_pdf("paginas/ReportUnit.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportUnitsXLSX(TemplateView):
    def get(self, request, user):
        user = int(user)
        # *? administrador
        if user == 0:
            rows = Unit.objects.all()
        # *? Empleado
        if user != 0:
            rows = Unit.objects.filter(user_id=user)

        units = rows
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Empleados"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Unidad"
        ws["E6"] = "Tipo de vehículo"
        ws["F6"] = "Kilometraje inicial"
        ws["G6"] = "Kilometraje Actual"
        ws["H6"] = "Marca"
        ws["I6"] = "Modelo"
        ws["J6"] = "Color"
        ws["K6"] = "Bastidor"
        ws["L6"] = "Motor"
        ws["M6"] = "Precio"
        ws["N6"] = "No. de permiso"
        ws["O6"] = "No. de placas"
        ws["P6"] = "Vigencia"
        ws["Q6"] = "Aseguradora"
        ws["R6"] = "No. de poliza"
        ws["S6"] = "Inicio de vigencia"
        ws["T6"] = "Final de vigencia"
        ws["U6"] = "No. de Tarjeta de circulación"
        ws["V6"] = "No. de Identificación"
        ws["W6"] = "Estado"

        cont = 7
        pos = 1

        for item in units:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.unit
            ws.cell(row=cont, column=5).value = item.vehicle_type.vehicle_type
            ws.cell(row=cont, column=6).value = item.initial_mileage
            ws.cell(row=cont, column=7).value = item.current_mileage
            ws.cell(row=cont, column=8).value = item.brand
            ws.cell(row=cont, column=9).value = item.model
            ws.cell(row=cont, column=10).value = item.color
            ws.cell(row=cont, column=11).value = item.frame
            ws.cell(row=cont, column=12).value = item.engine
            ws.cell(row=cont, column=13).value = item.price
            ws.cell(row=cont, column=14).value = item.permit_no
            ws.cell(row=cont, column=15).value = item.plate_no
            ws.cell(row=cont, column=16).value = item.validity
            ws.cell(row=cont, column=17).value = item.insurance_carrier
            ws.cell(row=cont, column=18).value = item.policy_no
            ws.cell(row=cont, column=19).value = item.start_validity
            ws.cell(row=cont, column=20).value = item.end_validity
            ws.cell(row=cont, column=21).value = item.circulation_card_no
            ws.cell(row=cont, column=22).value = item.identification_no
            ws.cell(row=cont, column=23).value = item.state
            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Unidades.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? COMBUSTIBLE
@api_view(["GET"])
def ListFuelType(request):
    fuel = FuelType.objects.all()
    serializer = FuelTypeSerializer(fuel, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def ListFueling(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("fuel_type" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        fuel_type = int(request.data["fuel_type"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (fuel_type == 0) and (user == 0):
            rows = Fueling.objects.filter(date__range=[fecha1, fecha2])
        if (fuel_type != 0) and (user == 0):
            rows = Fueling.objects.filter(
                date__range=[fecha1, fecha2], fuel_type_id=fuel_type
            )
        # *? Empleado
        if (fuel_type == 0) and (user != 0):
            rows = Fueling.objects.filter(date__range=[fecha1, fecha2], user_id=user)
        if (fuel_type != 0) and (user != 0):
            rows = Fueling.objects.filter(
                date__range=[fecha1, fecha2], user_id=user, fuel_type_id=fuel_type
            )

        serializer = FuelingListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateFueling(request):
    serializer = FuelingSerializer(data=request.data)
    data = request.data
    unit = Unit.objects.filter(id=data["unit"]).update(
        current_mileage=data["mileage"], hours_worked=data["hours"]
    )
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def DetailFueling(request, pk):
    fuel = Fueling.objects.get(id=pk)
    serializer = FuelingListSerializer(fuel, many=False)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteFueling(request, pk):
    fuel = Fueling.objects.get(id=pk)
    fuel.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["POST"])
def UpdateFueling(request, pk):
    fuel = Fueling.objects.get(id=pk)
    serializer = FuelingSerializer(instance=fuel, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ReportFuelLoadPDF(request, start_date, end_date, fuel_type, user):
    fecha1 = start_date
    fecha2 = end_date
    fuel_type = int(fuel_type)
    user = int(user)
    # *? administrador
    if user == 0:
        rows = Fueling.objects.filter(date__range=[fecha1, fecha2]).filter(
            fuel_type=fuel_type
        )
    # *? Empleado
    if user != 0:
        rows = (
            Fueling.objects.filter(date__range=[fecha1, fecha2])
            .filter(fuel_type=fuel_type)
            .filter(user_id=user)
        )

    data = {"fuels": rows}
    pdf = render_to_pdf("paginas/ReportFuelLoad.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportFuelLoadXLSX(TemplateView):
    def get(self, request, start_date, end_date, fuel_type, user):
        fecha1 = start_date
        fecha2 = end_date
        fuel_type = int(fuel_type)
        user = int(user)
        # *? administrador
        if user == 0:
            rows = Fueling.objects.filter(date__range=[fecha1, fecha2]).filter(
                fuel_type=fuel_type
            )
        # *? Empleado
        if user != 0:
            rows = (
                Fueling.objects.filter(date__range=[fecha1, fecha2])
                .filter(fuel_type=fuel_type)
                .filter(user_id=user)
            )

        fuels = rows
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Cargas de Combustibles"
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "Fecha"
        ws["E6"] = "Unidad"
        ws["F6"] = "Horas"
        ws["G6"] = "Litros"
        ws["H6"] = "Km/Lt"

        cont = 7
        pos = 1

        for item in fuels:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.date
            ws.cell(row=cont, column=5).value = item.unit.unit
            ws.cell(row=cont, column=6).value = item.hours
            ws.cell(row=cont, column=7).value = item.liters
            ws.cell(row=cont, column=8).value = item.km_liters

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_combustibles.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


@api_view(["GET"])
def ObtainLoads(request, pk):
    fuel = Fueling.objects.filter(unit_id=pk)
    serializer = FuelingListSerializer(fuel, many=True)
    return Response(serializer.data)


# *? DESCARGA DE COMBUSTIBLE


@api_view(["POST"])
def CreateFuelDump(request):
    serializer = FuelDumpSerializer(data=request.data)
    data = request.data
    # unit = Unit.objects.filter(id=data["unit"]).update(current_mileage = data["mileage"])
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateFuelDump(request, pk):
    row = FuelDump.objects.get(id=pk)
    serializer = FuelDumpSerializer(instance=row, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListFuelDump(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("fuel_type" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        fuel_type = int(request.data["fuel_type"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (fuel_type == 0) and (user == 0):
            rows = FuelDump.objects.filter(date__range=[fecha1, fecha2])
        if (fuel_type != 0) and (user == 0):
            rows = FuelDump.objects.filter(
                date__range=[fecha1, fecha2], fuel_type_id=fuel_type
            )
        # *? Empleado
        if (fuel_type == 0) and (user != 0):
            rows = FuelDump.objects.filter(date__range=[fecha1, fecha2], user_id=user)
        if (fuel_type != 0) and (user != 0):
            rows = FuelDump.objects.filter(
                date__range=[fecha1, fecha2], user_id=user, fuel_type_id=fuel_type
            )

        serializer = FuelDumpListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def DetailFuelDump(request, pk):
    fuel = FuelDump.objects.get(id=pk)
    serializer = FuelDumpListSerializer(fuel, many=False)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteFuelDump(request, pk):
    fuel = FuelDump.objects.get(id=pk)
    fuel.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? Requisición
@api_view(["POST"])
def Requisitions(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]

        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Requisition.objects.filter(date__range=[fecha1, fecha2])

        # *? Empleado
        if user != 0:
            rows = Requisition.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        serializer = RequisitionListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListRequisition(request):
    requisition = Requisition.objects.all()
    serializer = RequisitionListSerializer(requisition, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def ListAuthorizedRequisitions(request):
    requisition = Requisition.objects.filter(status="Autorizada")
    serializer = RequisitionListSerializer(requisition, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateRequisition(request):
    data = request.data
    new_requisition = Requisition.objects.create(
        date=data["date"],
        folio=data["folio"],
        applicant_id=data["applicant"],
        description=data["description"],
        observation=data["observation"],
        department_id=data["department"],
        status=data["status"],
        user_id=data["user"],
    )
    new_requisition.save()

    for product in data["products"]:
        new_product = ProductRequisition.objects.create(
            amount=product["amount"],
            unit=product["unit"],
            product=product["product"],
            code=data["code"],
        )
        new_product.save()
        product_obj = ProductRequisition.objects.filter(
            product=product["product"], code=data["code"]
        ).get(product=product["product"])
        new_requisition.products.add(product_obj)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["DELETE"])
def DeleteRequisition(request, pk):
    requisition = Requisition.objects.get(id=pk)
    requisition.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def RequisitionDetail(request, pk):
    requisition = Requisition.objects.get(id=pk)
    serializer = RequisitionListSerializer(requisition, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateStatusRequisition(request, pk):
    data = request.data
    requisitions = Requisition.objects.filter(id=pk).update(status=data["status"])
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": data}
    )
    # requisitions = Requisition.objects.get(id=pk)
    # serializer = RequisitionSerializer(instance=requisitions, data=request.data)
    # if requisitions:
    # serializer.save()
    # return Response({"message":"Registro actualizado satisfactoriamente!","status":200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["POST"])
def UploadFileQuotation(request):
    serializer = FileQuotationSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def DetailFileQuotation(request, pk):
    documents = FileQuotation.objects.filter(requisition_id=pk)
    serializer = FileQuotationSerializer(documents, many=True)
    return Response(serializer.data)


# *? CLIENTES
@api_view(["POST"])
def AccountStatus(request):
    if "customer" in request.data:
        customer = int(request.data["customer"])
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]

        suma1 = (
            CreditNote.objects.filter(
                customer_id=customer, date__range=[fecha1, fecha2]
            )
            .annotate(TotalNotas=Sum("total"))
            .values("TotalNotas")
        )
        rows = (
            BilledIncome.objects.filter(
                customer_id=customer,
                date__range=[fecha1, fecha2],
            )
            .annotate(SumaNotas=Subquery(suma1))
            .values("date", "invoice", "total", "SumNotas")
            .order_by("date", "invoice")
        )
        return Response(rows)
        # serializer = ListAccountStatusSerializer(rows, many=True)
        # return Response(serializer.data)


@api_view(["POST"])
def CustomersBalance(request):
    if "user" in request.data:
        user = int(request.data["user"])
        company = int(request.data["company"])
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        # return Response({"department->":department,"user->":user})
        t_peso = (
            BilledIncome.objects.filter(
                customer_id=OuterRef("pk"), date__range=[fecha1, fecha2]
            )
            .values("customer_id")
            .annotate(Total=Sum("net_weight"))
            .values("Total")
        )
        t_factura = (
            BilledIncome.objects.filter(
                customer_id=OuterRef("pk"), date__range=[fecha1, fecha2]
            )
            .values("customer_id")
            .annotate(Total=Sum("total"))
            .values("Total")
        )
        t_notas = (
            CreditNote.objects.filter(
                customer_id=OuterRef("pk"), date__range=[fecha1, fecha2]
            )
            .values("customer_id")
            .annotate(Total=Sum("total"))
            .values("Total")
        )
        t_ingreso = (
            Income.objects.filter(
                customer_id=OuterRef("pk"), date__range=[fecha1, fecha2]
            )
            .values("customer_id")
            .annotate(Total=Sum("amount"))
            .values("Total")
        )
        # return Response(t_notas)
        # *? administrador
        if (user == 0) and (company == 0):
            rows = (
                Customer.objects.all()
                .annotate(weight=Subquery(t_peso))
                .annotate(shopping=Subquery(t_factura))
                .annotate(creditnotes=Subquery(t_notas))
                .annotate(payments=Subquery(t_ingreso))
            ).values("id", "name", "weight", "shopping", "creditnotes", "payments")
        if (user == 0) and (company != 0):
            rows = (
                Customer.objects.filter(company_id=company)
                .annotate(weight=Subquery(t_peso))
                .annotate(shopping=Subquery(t_factura))
                .annotate(creditnotes=Subquery(t_notas))
                .annotate(payments=Subquery(t_ingreso))
            ).values("id", "name", "weight", "shopping", "creditnotes", "payments")

        # *? Empleado
        if (user != 0) and (company == 0):
            rows = (
                Customer.objects.filter(user_id=user)
                .annotate(weight=Subquery(t_peso))
                .annotate(shopping=Subquery(t_factura))
                .annotate(creditnotes=Subquery(t_notas))
                .annotate(payments=Subquery(t_ingreso))
            ).values("id", "name", "weight", "shopping", "creditnotes", "payments")
        if (user != 0) and (company != 0):
            rows = (
                Customer.objects.filter(user_id=user)
                .filter(company_id=company)
                .annotate(weight=Subquery(t_peso))
                .annotate(shopping=Subquery(t_factura))
                .annotate(creditnotes=Subquery(t_notas))
                .annotate(payments=Subquery(t_ingreso))
            ).values("id", "name", "weight", "shopping", "creditnotes", "payments")

        # serializer = CustomerSerializer(rows, many=True)
        # return Response(serializer.data)
        return Response(rows)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def Customers(request):
    if "user" in request.data:
        user = int(request.data["user"])
        company = int(request.data["company"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (user == 0) and (company == 0):
            rows = Customer.objects.all()
        if (user == 0) and (company != 0):
            rows = Customer.objects.filter(company_id=company)

        # *? Empleado
        if (user != 0) and (company == 0):
            rows = Customer.objects.filter(user_id=user)
        if (user != 0) and (company != 0):
            rows = Customer.objects.filter(user_id=user).filter(company_id=company)

        serializer = CustomerSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListCustomer(request):
    customer = Customer.objects.all()
    serializer = CustomerSerializer(customer, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def CustomerList(request, pk):
    customer = Customer.objects.filter(user_id=pk)
    serializer = CustomerSerializer(customer, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def CustomerDetail(request, pk):
    customer = Customer.objects.get(id=pk)
    serializer = CustomerSerializer(customer, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateCustomer(request):
    serializer = CustomerSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def UpdateCustomer(request, pk):
    customer = Customer.objects.get(id=pk)
    serializer = CustomerSerializer(instance=customer, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteCustomer(request, pk):
    customer = Customer.objects.get(id=pk)
    customer.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["POST"])
def UpdatePhotoCustomer(request, pk):
    data = request.data
    customer = Customer.objects.get(id=pk)
    customer.imagen = data["imagen"]
    customer.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


# class ReportCustomersPDF(View):
@api_view(["GET"])
def ReportCustomersPDF(request, user):
    user = int(user)
    # return Response({"department->":department,"user->":user})
    # *? administrador
    if user == 0:
        rows = Customer.objects.all()

    # *? Empleado
    if user != 0:
        rows = Customer.objects.filter(user_id=user)

    customers = rows
    # customers = Customer.objects.all()
    # serializer = EmployeeListSerializer(employees, many=True)
    data = {"customers": customers}
    pdf = render_to_pdf("paginas/ReportCustomers.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportCustomersXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        customers = Customer.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")

        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:I2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:I3")
        ws["C5"] = "Reporte de Empleados"
        ws.merge_cells("C5:I5")

        ws["C6"] = "ID"
        ws["D6"] = "Nombre"
        ws["E6"] = "RFC"
        ws["F6"] = "Teléfono"
        ws["G6"] = "Correo electrónico"
        ws["H6"] = "Representante"
        ws["I6"] = "Codigo Postal"
        ws["J6"] = "Estado"
        ws["K6"] = "Ciudad"
        ws["L6"] = "Localidad"
        ws["M6"] = "Colonia"
        ws["N6"] = "Calle"
        ws["O6"] = "No. Interior"
        ws["P6"] = "No. Exterior"

        cont = 7
        pos = 1

        for customer in customers:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = customer.name
            ws.cell(row=cont, column=5).value = customer.rfc
            ws.cell(row=cont, column=6).value = customer.phone
            ws.cell(row=cont, column=7).value = customer.email
            ws.cell(row=cont, column=8).value = customer.representative
            ws.cell(row=cont, column=9).value = customer.postal_code
            ws.cell(row=cont, column=10).value = customer.country
            ws.cell(row=cont, column=11).value = customer.state
            ws.cell(row=cont, column=12).value = customer.city
            ws.cell(row=cont, column=13).value = customer.location
            ws.cell(row=cont, column=14).value = customer.suburb
            ws.cell(row=cont, column=15).value = customer.int_no
            ws.cell(row=cont, column=16).value = customer.ext_no
            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Clientes.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? ORDEN DE PEDIDO
@api_view(["POST"])
def PurchaseOrders(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = PurchaseOrder.objects.filter(date__range=[fecha1, fecha2]).order_by(
                "purchase_order"
            )

        if (customer != 0) and (user == 0):
            rows = (
                PurchaseOrder.objects.filter(business_name_id=customer)
                .filter(date__range=[fecha1, fecha2])
                .order_by("purchase_order")
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = (
                PurchaseOrder.objects.filter(user_id=user)
                .filter(date__range=[fecha1, fecha2])
                .order_by("purchase_order")
            )

        if (customer != 0) and (user != 0):
            rows = (
                PurchaseOrder.objects.filter(user_id=user)
                .filter(business_name_id=customer)
                .filter(date__range=[fecha1, fecha2])
            ).order_by("purchase_order")

        serializer = PurchaseOrderListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListPurchaseOrders(request):
    users = PurchaseOrder.objects.all()
    serializer = ListPurchaseOrdersSerializers(users, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreatePurchaseOrder(request):
    data = request.data
    new_creditnote = PurchaseOrder.objects.create(
        date=data["date"],
        purchase_order=data["purchase_order"],
        date_delivery=data["date_delivery"],
        business_name_id=data["business_name"],
        order_number=data["order_number"],
        place_delivery=data["place_delivery"],
        sale_condition=data["sale_condition"],
        invoice=data["invoice"],
        observation=data["observation"],
        auxiliary_sales_id=data["auxiliary_sales"],
        storekeeper_id=data["storekeeper"],
        qa_id=data["qa"],
        status=data["status"],
        voucher=data["voucher"],
        # iva=data['iva'],
        # subtotal=data['subtotal'],
        total=data["total"],
        user_id=data["user"],
    )
    new_creditnote.save()

    for product in data["products"]:
        new_product = ProductPO.objects.create(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
            code=data["code"],
        )
        new_product.save()
        product_obj = ProductPO.objects.get(id=new_product.id)
        # product_obj = ProductPO.objects.filter(amount=product["amount"], unit=product["unit"], presentation_id=product["presentation"], product_id=product["product"], price=product["price"], subtotal = product["subtotal"], code=data["code"]).get(code=data["code"])
        new_creditnote.products.add(product_obj)

    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["POST"])
def UpdatePurchaseOrder(request, pk):
    data = request.data
    PurchaseOrder.objects.filter(id=pk).update(
        date=data["date"],
        purchase_order=data["purchase_order"],
        date_delivery=data["date_delivery"],
        business_name_id=data["business_name"],
        order_number=data["order_number"],
        place_delivery=data["place_delivery"],
        sale_condition=data["sale_condition"],
        invoice=data["invoice"],
        observation=data["observation"],
        auxiliary_sales_id=data["auxiliary_sales"],
        storekeeper_id=data["storekeeper"],
        qa_id=data["qa"],
        status=data["status"],
        voucher=data["voucher"],
        # iva=data['iva'],
        # subtotal=data['subtotal'],
        total=data["total"],
        user_id=data["user"],
    )
    for product in data["rows"]:
        new_product = ProductPO.objects.filter(id=product["id"]).update(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
        )
        # new_product.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["DELETE"])
def DeletePurchaseOrder(request, pk):
    orden = PurchaseOrder.objects.get(id=pk)
    items = orden.products.all()
    for item in items:
        producto = ProductPO.objects.get(id=item.id)
        producto.delete()
    orden.delete()
    return Response("Registro eliminado satisfactoriamente")
    # return Response({"filas":rows.products})
    # rows.delete()
    # return Response('Registro eliminado satisfactoriamente!')


@api_view(["GET"])
def DetailPurchaseOrder(request, pk):
    row = PurchaseOrder.objects.get(id=pk)
    serializer = PurchaseOrderListSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["GET"])
def ReportPurchaseOrderPDF(request, pk):
    pk = int(pk)
    row = PurchaseOrder.objects.get(id=pk)

    data = {"orden": row}
    pdf = render_to_pdf("paginas/PurchaseOrder.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


# *? NOTAS DE CREDITO
@api_view(["POST"])
def CreditNotes(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = CreditNote.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = CreditNote.objects.filter(business_name_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = CreditNote.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                CreditNote.objects.filter(user_id=user)
                .filter(business_name_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = CreditNoteListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def CreditNoteList(request):
    creditnote = CreditNote.objects.all()
    serializer = CreditNoteListSerializer(creditnote, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteCreditNote(request, pk):
    creditnote = CreditNote.objects.get(id=pk)
    items = creditnote.products.all()
    for item in items:
        producto = ProductCN.objects.get(id=item.id)
        producto.delete()
    creditnote.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def CreditNoteDetail(request, pk):
    creditnote = CreditNote.objects.get(id=pk)
    serializer = CreditNoteListSerializer(creditnote, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateCreditNote(request):
    data = request.data
    new_creditnote = CreditNote.objects.create(
        date=data["date"],
        no_creditnote=data["no_creditnote"],
        purchase_order=data["purchase_order"],
        date_delivery=data["date_delivery"],
        customer_id=data["business_name"],
        order_number=data["order_number"],
        place_delivery=data["place_delivery"],
        sale_condition=data["sale_condition"],
        invoice_id=data["invoice"],
        observation=data["observation"],
        auxiliary_sales_id=data["auxiliary_sales"],
        storekeeper_id=data["storekeeper"],
        qa_id=data["qa"],
        # iva=data['iva'],
        # subtotal=data['subtotal'],
        total=data["total"],
        user_id=data["user"],
    )
    new_creditnote.save()

    for product in data["products"]:
        new_product = ProductCN.objects.create(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
            code=data["code"],
        )
        new_product.save()
        product_obj = ProductCN.objects.get(id=new_product.id)
        # product_obj = ProductCN.objects.filter(amount=product["amount"], unit=product["unit"], presentation_id=product["presentation"], product_id=product["product"], price=product["price"], subtotal = product["subtotal"], code=data["code"]).get(code=data["code"])
        new_creditnote.products.add(product_obj)

    # serializer = RequisitionSerializer(new_creditnote)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["GET"])
def DetailCreditNote(request, pk):
    row = CreditNote.objects.get(id=pk)
    serializer = CreditNoteListSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateCreditNote(request, pk):
    data = request.data
    CreditNote.objects.filter(id=pk).update(
        date=data["date"],
        no_creditnote=data["no_creditnote"],
        purchase_order=data["purchase_order"],
        date_delivery=data["date_delivery"],
        customer_id=data["business_name"],
        order_number=data["order_number"],
        place_delivery=data["place_delivery"],
        sale_condition=data["sale_condition"],
        invoice_id=data["invoice"],
        observation=data["observation"],
        auxiliary_sales_id=data["auxiliary_sales"],
        storekeeper_id=data["storekeeper"],
        qa_id=data["qa"],
        # iva=data['iva'],
        # subtotal=data['subtotal'],
        total=data["total"],
        user_id=data["user"],
    )
    for product in data["rows"]:
        new_product = ProductCN.objects.filter(id=product["id"]).update(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
        )
        # new_product.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["GET"])
def ReportCreditNotePDF(request, pk):
    pk = int(pk)
    row = CreditNote.objects.get(id=pk)

    data = {"orden": row}
    pdf = render_to_pdf("paginas/CreditNote.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


# *? COTIZACIONES
@api_view(["POST"])
def Quotes(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = Quotation.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = Quotation.objects.filter(customer_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = Quotation.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                Quotation.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListQuotationSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListQuotation(request):
    quotation = Quotation.objects.all()
    serializer = ListQuotationSerializer(quotation, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteQuotation(request, pk):
    quotation = Quotation.objects.get(id=pk)
    items = quotation.products.all()
    for item in items:
        producto = ProductQuotation.objects.get(id=item.id)
        producto.delete()
    quotation.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailQuotation(request, pk):
    quotation = Quotation.objects.get(id=pk)
    serializer = ListQuotationSerializer(quotation, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateQuotation(request):
    data = request.data
    new_quotation = Quotation.objects.create(
        date=data["date"],
        invoice=data["invoice"],
        period=data["period"],
        source=data["source"],
        attention=data["attention"],
        customer_id=data["customer"],
        introduction=data["introduction"],
        # iva=data['iva'],
        # subtotal=data['subtotal'],
        total=data["total"],
        generator_id=data["generator"],
        authorizer_id=data["authorizer"],
        comment=data["comment"],
        user_id=data["user"],
    )
    new_quotation.save()

    for product in data["products"]:
        new_product = ProductQuotation.objects.create(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
            code=data["code"],
        )
        new_product.save()
        # product_obj = ProductQuotation.objects.filter(amount=product["amount"], unit=product["unit"], presentation_id=product["presentation"], product_id=product["product"], price=product["price"], subtotal = product["subtotal"], code=data["code"]).get(code=data["code"])
        product_obj = ProductQuotation.objects.get(id=new_product.id)
        new_quotation.products.add(product_obj)

    # serializer = RequisitionSerializer(new_creditnote)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["POST"])
def UpdateQuotation(request, pk):
    data = request.data
    Quotation.objects.filter(id=pk).update(
        date=data["date"],
        invoice=data["invoice"],
        period=data["period"],
        source=data["source"],
        attention=data["attention"],
        customer_id=data["customer"],
        introduction=data["introduction"],
        # iva=data['iva'],
        # subtotal=data['subtotal'],
        total=data["total"],
        generator_id=data["generator"],
        authorizer_id=data["authorizer"],
        comment=data["comment"],
    )
    for product in data["rows"]:
        new_product = ProductQuotation.objects.filter(id=product["id"]).update(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
        )
        # new_product.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["GET"])
def ReportQuotationPDF(request, pk):
    pk = int(pk)
    row = Quotation.objects.get(id=pk)

    data = {"orden": row}
    pdf = render_to_pdf("paginas/Quotation.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


# *? ALMACÉN
@api_view(["GET"])
def WarehouseList(request):
    warehouse = Warehouse.objects.all()
    serializer = WarehouseListSerializer(warehouse, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateWarehouse(request):
    serializer = WarehouseSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? PRESENTACIÓN ALMACÉN
@api_view(["GET"])
def PresentationList(request):
    presentation = Presentation.objects.all()
    serializer = PresentationSerializer(presentation, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreatePresentation(request):
    serializer = PresentationSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeletePresentation(request, pk):
    presentation = Presentation.objects.get(id=pk)
    presentation.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def PresentationDetail(request, pk):
    presentation = Presentation.objects.get(id=pk)
    serializer = PresentationSerializer(presentation, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdatePresentation(request, pk):
    presentation = Presentation.objects.get(id=pk)
    serializer = PresentationSerializer(instance=presentation, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


class ReportPresentationsPDF(View):
    def get(self, request, *args, **kwargs):
        presentations = Presentation.objects.all()
        # serializer = EmployeeListSerializer(employees, many=True)
        data = {"presentations": presentations}
        pdf = render_to_pdf("paginas/ReportPresentations.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportPresentationsXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        presentations = Presentation.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")

        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:I2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:I3")
        ws["C5"] = "Reporte de Presentaciones"
        ws.merge_cells("C5:I5")

        ws["C6"] = "ID"
        ws["D6"] = "Nombre"

        cont = 7
        pos = 1

        for item in presentations:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.name

            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Presentaciones.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? PRODUCTOS PRINCIPALES
@api_view(["POST"])
def CreateMainProduct(request):
    serializer = MainProductSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListMainProduct(request):
    product = MainProduct.objects.all()
    serializer = MainProductSerializer(product, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def ListMainPresentation(request):
    product = MainPresentation.objects.all()
    serializer = MainPresentationSerializer(product, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteMainProduct(request, pk):
    product = MainProduct.objects.get(id=pk)
    product.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailMainProduct(request, pk):
    product = MainProduct.objects.get(id=pk)
    serializer = MainProductSerializer(product, many=False)
    return Response(serializer.data)


# *? PRODUCTOS ALMACÉN
@api_view(["GET"])
def ListProduct(request):
    product = Product.objects.all()
    serializer = ProductSerializer(product, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateProduct(request):
    serializer = ProductSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteProduct(request, pk):
    product = Product.objects.get(id=pk)
    product.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailProduct(request, pk):
    product = Product.objects.get(id=pk)
    serializer = ProductSerializer(product, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateProduct(request, pk):
    product = Product.objects.get(id=pk)
    serializer = ProductSerializer(instance=product, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


class ReportProductsPDF(View):
    def get(self, request, *args, **kwargs):
        products = Product.objects.all()
        # serializer = EmployeeListSerializer(employees, many=True)
        data = {"products": products}
        pdf = render_to_pdf("paginas/ReportProducts.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportProductsXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        products = Product.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")

        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:I2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:I3")
        ws["C5"] = "Reporte de Productos"
        ws.merge_cells("C5:I5")

        ws["C6"] = "ID"
        ws["D6"] = "Nombre"

        cont = 7
        pos = 1

        for item in products:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.name

            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Productos.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? REPORTE DE REPASO DE ENTRADAS


@api_view(["POST"])
def CreateTicketReview(request):
    data = request.data
    new_ticket = Ticketreview.objects.create(
        date=data["date"],
        variety=data["variety"],
        total=data["total"],
        observation=data["observation"],
    )
    new_ticket.save()

    for row in data["rows"]:
        new_product = Rowticketreview.objects.create(
            product_id=row["product"],
            bigbags_id=row["bigbags"],
            kgsxbigbags=row["kgsxbigbags"],
            packages_id=row["packages"],
            kgsxpackages=row["kgsxpackages"],
            subtotal=row["subtotal"],
            code=data["code"],
        )
        new_product.save()

        row_obj = Rowticketreview.objects.filter(
            product_id=row["product"], code=data["code"]
        ).get(product_id=row["product"])
        new_ticket.rows.add(row_obj)

    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["GET"])
def TicketReviewList(request):
    ticketreview = Ticketreview.objects.all()
    serializer = TicketreviewListSerializer(ticketreview, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def TicketReviewDetail(request, pk):
    ticketreview = Ticketreview.objects.get(id=pk)
    serializer = TicketreviewListSerializer(ticketreview, many=False)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteTicketReview(request, pk):
    ticketreview = Ticketreview.objects.get(id=pk)
    ticketreview.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? SALIDAS
@api_view(["POST"])
def Outputs(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = Output.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = Output.objects.filter(customer_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = Output.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                Output.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = OutputListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateOutput(request):
    data = request.data
    new_output = Output.objects.create(
        user_id=data["user"],
        company_id=data["company"],
        date=data["date"],
        invoice=data["invoice"],
        customer_id=data["customer"],
        destiny=data["destiny"],
        gross_weight=data["gross_weight"],
        tare=data["tare"],
        net_weight=data["net_weight"],
        driver_id=data["driver"],
        plate_id=data["plate"],
        payment_method=data["payment_method"],
        bank_account_id=data["bank_account"],
        # subtotal=data['subtotal'],
        # iva=data['iva'],
        total=data["total"],
        delivered_id=data["delivered"],
        received=data["received"],
        comment=data["comment"],
    )
    new_output.save()
    no = 1
    for row in data["products"]:
        new_product = OutputProduct.objects.create(
            no=no,
            amount=row["amount"],
            unit=row["unit"],
            presentation_id=row["presentation"],
            product_id=row["product"],
            price=row["price"],
            subtotal=row["subtotal"],
            code=data["code"],
        )
        new_product.save()
        row_obj = OutputProduct.objects.get(id=new_product.id)
        new_output.products.add(row_obj)
        no += 1

    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["POST"])
def UpdateOutput(request, pk):
    data = request.data
    Output.objects.filter(id=pk).update(
        date=data["date"],
        company_id=data["company"],
        invoice=data["invoice"],
        customer_id=data["customer"],
        destiny=data["destiny"],
        gross_weight=data["gross_weight"],
        tare=data["tare"],
        net_weight=data["net_weight"],
        driver_id=data["driver"],
        plate_id=data["plate"],
        payment_method=data["payment_method"],
        bank_account_id=data["bank_account"],
        # subtotal=data['subtotal'],
        # iva=data['iva'],
        total=data["total"],
        delivered_id=data["delivered"],
        received=data["received"],
        comment=data["comment"],
    )
    for product in data["rows"]:
        new_product = OutputProduct.objects.filter(id=product["id"]).update(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            price=product["price"],
            subtotal=product["subtotal"],
        )
        # new_product.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["POST"])
def OutputList(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    company = data["company"]
    output = Output.objects.filter(date__range=[fecha1, fecha2]).filter(
        company_id=company
    )
    serializer = OutputListSerializer(output, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def OutputDetail(request, pk):
    output = Output.objects.get(id=pk)
    serializer = OutputListSerializer(output, many=False)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteOutput(request, pk):
    output = Output.objects.get(id=pk)
    items = output.products.all()
    for item in items:
        producto = OutputProduct.objects.get(id=item.id)
        producto.delete()
    output.delete()
    return Response("Registro eliminado satisfactoriamente!")


# ? BITACORA
@api_view(["POST"])
def Binnacles(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Binnacle.objects.filter(fecha__range=[fecha1, fecha2])

        # *? Empleado
        if user != 0:
            rows = Binnacle.objects.filter(fecha__range=[fecha1, fecha2], user_id=user)

        serializer = BinnacleSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListBinnacle(request, pk):
    binnacle = Binnacle.objects.filter(user_id=pk)
    serializer = BinnacleSerializer(binnacle, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateBinnacle(request):
    serializer = BinnacleSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteBinnacle(request, pk):
    binnacle = Binnacle.objects.get(id=pk)
    binnacle.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailBinnacle(request, pk):
    binnacle = Binnacle.objects.get(id=pk)
    serializer = BinnacleSerializer(binnacle, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateBinnacle(request, pk):
    binnacle = Binnacle.objects.get(id=pk)
    serializer = BinnacleSerializer(instance=binnacle, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ReportBinnaclesPDF(request, start_date, end_date, user):
    fecha1 = start_date
    fecha2 = end_date
    user = int(user)
    # return Response({"department->":department,"user->":user})
    # *? administrador
    if user == 0:
        rows = Binnacle.objects.filter(fecha__range=[fecha1, fecha2])

    # *? Empleado
    if user != 0:
        rows = Binnacle.objects.filter(user_id=user).filter(
            fecha__range=[fecha1, fecha2]
        )

    data = {"binnacles": rows}
    pdf = render_to_pdf("paginas/ReportBinnacles.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportBinnaclesXLSX(TemplateView):
    def get(self, request, start_date, end_date, user):
        fecha1 = start_date
        fecha2 = end_date
        user = int(user)
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Binnacle.objects.filter(fecha__range=[fecha1, fecha2])
        # *? Empleado
        if user != 0:
            rows = Binnacle.objects.filter(user_id=user).filter(
                fecha__range=[fecha1, fecha2]
            )
        # binnacles = Binnacle.objects.all()
        binnacles = rows
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de Bitacoras"
        ws.merge_cells("C5:J5")

        borde_negro = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
        )

        ws["C6"] = "ID"
        ws.column_dimensions["C"].width = 20
        ws["C6"].border = borde_negro

        ws["D6"] = "Fecha"
        ws.column_dimensions["D"].width = 20
        ws["D6"].border = borde_negro

        ws["E6"] = "Labores"
        ws.column_dimensions["E"].width = 20
        ws["E6"].border = borde_negro

        ws["F6"] = "Lotes"
        ws.column_dimensions["F"].width = 20
        ws["F6"].border = borde_negro

        ws["G6"] = "HA"
        ws.column_dimensions["G"].width = 20
        ws["G6"].border = borde_negro

        ws["H6"] = "Faena"
        ws.column_dimensions["H"].width = 20
        ws["H6"].border = borde_negro

        ws["I6"] = "Total aplicado"
        ws.column_dimensions["I"].width = 20
        ws["I6"].border = borde_negro

        ws["J6"] = "medida"
        ws.column_dimensions["J"].width = 20
        ws["J6"].border = borde_negro

        ws["K6"] = "Acurate"
        ws.column_dimensions["K"].width = 20
        ws["K6"].border = borde_negro

        ws["L6"] = "Total aplicado"
        ws.column_dimensions["L"].width = 20
        ws["L6"].border = borde_negro

        ws["M6"] = "medida"
        ws.column_dimensions["M"].width = 20
        ws["M6"].border = borde_negro

        ws["N6"] = "Prolux"
        ws.column_dimensions["N"].width = 20
        ws["N6"].border = borde_negro

        ws["O6"] = "Total aplicado"
        ws.column_dimensions["O"].width = 20
        ws["O6"].border = borde_negro

        ws["P6"] = "medida"
        ws.column_dimensions["P"].width = 20
        ws["P6"].border = borde_negro

        ws["Q6"] = "Aderente"
        ws.column_dimensions["Q"].width = 20
        ws["Q6"].border = borde_negro

        ws["R6"] = "Total aplicado"
        ws.column_dimensions["R"].width = 20
        ws["R6"].border = borde_negro

        ws["S6"] = "medida"
        ws.column_dimensions["S"].width = 20
        ws["S6"].border = borde_negro

        ws["T6"] = "Sulfato de amonio"
        ws.column_dimensions["T"].width = 20
        ws["T6"].border = borde_negro

        ws["U6"] = "Total aplicado"
        ws.column_dimensions["U"].width = 20
        ws["U6"].border = borde_negro

        ws["V6"] = "medida"
        ws.column_dimensions["V"].width = 20
        ws["V6"].border = borde_negro

        ws["W6"] = "Nomolt"
        ws.column_dimensions["W"].width = 20
        ws["W6"].border = borde_negro

        ws["X6"] = "Total aplicado"
        ws.column_dimensions["X"].width = 20
        ws["X6"].border = borde_negro

        ws["Y6"] = "medida"
        ws.column_dimensions["Y"].width = 20
        ws["Y6"].border = borde_negro

        ws["Z6"] = "Metarrizium"
        ws.column_dimensions["Z"].width = 20
        ws["Z6"].border = borde_negro

        ws["AA6"] = "Total aplicado"
        ws.column_dimensions["AA"].width = 20
        ws["AA6"].border = borde_negro

        ws["AB6"] = "medida"
        ws.column_dimensions["AB"].width = 20
        ws["AB6"].border = borde_negro

        ws["AC6"] = "Beauberian"
        ws.column_dimensions["AC"].width = 20
        ws["AC6"].border = borde_negro

        ws["AD6"] = "Total aplicado"
        ws.column_dimensions["AD"].width = 20
        ws["AD6"].border = borde_negro

        ws["AE6"] = "medida"
        ws.column_dimensions["AE"].width = 20
        ws["AE6"].border = borde_negro

        ws["AF6"] = "Abland"
        ws.column_dimensions["AF"].width = 20
        ws["AF6"].border = borde_negro

        ws["AG6"] = "Total aplicado"
        ws.column_dimensions["AG"].width = 20
        ws["AG6"].border = borde_negro

        ws["AH6"] = "medida"
        ws.column_dimensions["AH"].width = 20
        ws["AH6"].border = borde_negro

        ws["AI6"] = "aura"
        ws.column_dimensions["AI"].width = 20
        ws["AI6"].border = borde_negro

        ws["AJ6"] = "Total aplicado"
        ws.column_dimensions["AJ"].width = 20
        ws["AJ6"].border = borde_negro

        ws["AK6"] = "medida"
        ws.column_dimensions["AK"].width = 20
        ws["AK6"].border = borde_negro

        ws["AL6"] = "Urea"
        ws.column_dimensions["AL"].width = 20
        ws["AL6"].border = borde_negro

        ws["AM6"] = "Total aplicado"
        ws.column_dimensions["AM"].width = 20
        ws["AM6"].border = borde_negro

        ws["AN6"] = "medida"
        ws.column_dimensions["AN"].width = 20
        ws["AN6"].border = borde_negro

        ws["AO6"] = "Dash"
        ws.column_dimensions["AO"].width = 20
        ws["AO6"].border = borde_negro

        ws["AP6"] = "Total aplicado"
        ws.column_dimensions["AP"].width = 20
        ws["AP6"].border = borde_negro

        ws["AQ6"] = "medida"
        ws.column_dimensions["AQ"].width = 20
        ws["AQ6"].border = borde_negro

        ws["AR6"] = "Cipermetrina"
        ws.column_dimensions["AR"].width = 20
        ws["AR6"].border = borde_negro

        ws["AS6"] = "Total aplicado"
        ws.column_dimensions["AS"].width = 20
        ws["AS6"].border = borde_negro

        ws["AT6"] = "medida"
        ws.column_dimensions["AT"].width = 20
        ws["AT6"].border = borde_negro

        cont = 7
        pos = 1
        for item in binnacles:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=3).border = borde_negro
            ws.cell(row=cont, column=4).value = item.fecha
            ws.cell(row=cont, column=4).border = borde_negro
            ws.cell(row=cont, column=5).value = item.labores
            ws.cell(row=cont, column=5).border = borde_negro
            ws.cell(row=cont, column=6).value = item.lotes
            ws.cell(row=cont, column=6).border = borde_negro
            ws.cell(row=cont, column=7).value = item.ha
            ws.cell(row=cont, column=7).border = borde_negro
            ws.cell(row=cont, column=8).value = item.faena
            ws.cell(row=cont, column=8).border = borde_negro
            ws.cell(row=cont, column=9).value = item.ta_ha_faena
            ws.cell(row=cont, column=9).border = borde_negro
            ws.cell(row=cont, column=10).value = item.medida_faena
            ws.cell(row=cont, column=10).border = borde_negro
            ws.cell(row=cont, column=11).value = item.acurate
            ws.cell(row=cont, column=11).border = borde_negro
            ws.cell(row=cont, column=12).value = item.ta_ha_acurate
            ws.cell(row=cont, column=12).border = borde_negro
            ws.cell(row=cont, column=13).value = item.medida_acurate
            ws.cell(row=cont, column=13).border = borde_negro
            ws.cell(row=cont, column=14).value = item.prolux
            ws.cell(row=cont, column=14).border = borde_negro
            ws.cell(row=cont, column=15).value = item.ta_ha_prolux
            ws.cell(row=cont, column=15).border = borde_negro
            ws.cell(row=cont, column=16).value = item.medida_prolux
            ws.cell(row=cont, column=16).border = borde_negro
            ws.cell(row=cont, column=17).value = item.aderente
            ws.cell(row=cont, column=17).border = borde_negro
            ws.cell(row=cont, column=18).value = item.ta_ha_aderente
            ws.cell(row=cont, column=18).border = borde_negro
            ws.cell(row=cont, column=19).value = item.medida_aderente
            ws.cell(row=cont, column=19).border = borde_negro
            ws.cell(row=cont, column=20).value = item.sulfato
            ws.cell(row=cont, column=20).border = borde_negro
            ws.cell(row=cont, column=21).value = item.ta_ha_sulfato
            ws.cell(row=cont, column=21).border = borde_negro
            ws.cell(row=cont, column=22).value = item.medida_sulfato
            ws.cell(row=cont, column=22).border = borde_negro
            ws.cell(row=cont, column=23).value = item.nomolt
            ws.cell(row=cont, column=23).border = borde_negro
            ws.cell(row=cont, column=24).value = item.ta_ha_nomolt
            ws.cell(row=cont, column=24).border = borde_negro
            ws.cell(row=cont, column=25).value = item.medida_nomolt
            ws.cell(row=cont, column=25).border = borde_negro
            ws.cell(row=cont, column=26).value = item.metarrizium
            ws.cell(row=cont, column=26).border = borde_negro
            ws.cell(row=cont, column=27).value = item.ta_ha_metarrizium
            ws.cell(row=cont, column=27).border = borde_negro
            ws.cell(row=cont, column=28).value = item.medida_metarrizium
            ws.cell(row=cont, column=28).border = borde_negro
            ws.cell(row=cont, column=29).value = item.beauberian
            ws.cell(row=cont, column=29).border = borde_negro
            ws.cell(row=cont, column=30).value = item.ta_ha_beauberian
            ws.cell(row=cont, column=30).border = borde_negro
            ws.cell(row=cont, column=31).value = item.medida_beauberian
            ws.cell(row=cont, column=31).border = borde_negro
            ws.cell(row=cont, column=32).value = item.abland
            ws.cell(row=cont, column=32).border = borde_negro
            ws.cell(row=cont, column=33).value = item.ta_ha_abland
            ws.cell(row=cont, column=33).border = borde_negro
            ws.cell(row=cont, column=34).value = item.medida_abland
            ws.cell(row=cont, column=34).border = borde_negro
            ws.cell(row=cont, column=35).value = item.aura
            ws.cell(row=cont, column=35).border = borde_negro
            ws.cell(row=cont, column=36).value = item.ta_ha_aura
            ws.cell(row=cont, column=36).border = borde_negro
            ws.cell(row=cont, column=37).value = item.medida_aura
            ws.cell(row=cont, column=37).border = borde_negro
            ws.cell(row=cont, column=38).value = item.urea
            ws.cell(row=cont, column=38).border = borde_negro
            ws.cell(row=cont, column=39).value = item.ta_ha_urea
            ws.cell(row=cont, column=39).border = borde_negro
            ws.cell(row=cont, column=40).value = item.medida_urea
            ws.cell(row=cont, column=40).border = borde_negro
            ws.cell(row=cont, column=41).value = item.dash
            ws.cell(row=cont, column=41).border = borde_negro
            ws.cell(row=cont, column=42).value = item.ta_ha_dash
            ws.cell(row=cont, column=42).border = borde_negro
            ws.cell(row=cont, column=43).value = item.medida_dash
            ws.cell(row=cont, column=43).border = borde_negro
            ws.cell(row=cont, column=44).value = item.cipermetrina
            ws.cell(row=cont, column=44).border = borde_negro
            ws.cell(row=cont, column=45).value = item.ta_ha_cipermetrina
            ws.cell(row=cont, column=45).border = borde_negro
            ws.cell(row=cont, column=46).value = item.medida_cipermetrina
            ws.cell(row=cont, column=46).border = borde_negro

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_bitacoras.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# ? PARCELAS
@api_view(["POST"])
def Parcels(request):
    if "user" in request.data:
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Parcel.objects.all()

        # *? Empleado
        if user != 0:
            rows = Parcel.objects.filter(user_id=user)

        serializer = ParcelListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListParcel(request):
    parcel = Parcel.objects.all()
    serializer = ParcelListSerializer(parcel, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateParcel(request):
    serializer = ParcelSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteParcel(request, pk):
    parcel = Parcel.objects.get(id=pk)
    parcel.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailParcel(request, pk):
    parcel = Parcel.objects.get(id=pk)
    serializer = ParcelSerializer(parcel, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateParcel(request, pk):
    parcel = Parcel.objects.get(id=pk)
    serializer = ParcelSerializer(instance=parcel, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# class ReportParcelsPDF(View):
@api_view(["GET"])
def ReportParcelsPDF(request, user):
    user = int(user)
    # *? administrador
    if user == 0:
        rows = Parcel.objects.all()
    # *? Empleado
    if user != 0:
        rows = Parcel.objects.filter(user_id=user)
    # parcels = Parcel.objects.all()
    # serializer = EmployeeListSerializer(employees, many=True)
    data = {"parcels": rows}
    pdf = render_to_pdf("paginas/ReportParcel.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportParcelsXLSX(TemplateView):
    def get(self, request, user):
        user = int(user)
        # *? administrador
        if user == 0:
            rows = Parcel.objects.all()
        # *? Empleado
        if user != 0:
            rows = Parcel.objects.filter(user_id=user)

        parcels = rows
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")

        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:I2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:I3")
        ws["C5"] = "Reporte de Parcelas"
        ws.merge_cells("C5:I5")

        ws["C6"] = "ID"
        ws["D6"] = "No. Parcela"
        ws["E6"] = "Georeferencia"
        ws["F6"] = "Hectareas"
        ws["G6"] = "Certificado"
        ws["H6"] = "RAN"
        ws["I6"] = "Propietario"
        ws["J6"] = "Ubicación"
        ws["K6"] = "Comentario"

        cont = 7
        pos = 1

        for item in parcels:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.no_parcel
            ws.cell(row=cont, column=5).value = item.georeferences
            ws.cell(row=cont, column=6).value = item.hectares
            ws.cell(row=cont, column=7).value = item.certificate
            ws.cell(row=cont, column=8).value = item.ran
            ws.cell(row=cont, column=9).value = item.owner
            ws.cell(row=cont, column=10).value = item.location
            ws.cell(row=cont, column=11).value = item.comment

            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Sociedades.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# ? RENTA DE PREDIOS
@api_view(["POST"])
def ListLandRent(request):
    if (
        ("rent_year" in request.data)
        and ("reason_rent" in request.data)
        and ("user" in request.data)
    ):
        rent_year = request.data["rent_year"]
        reason_rent = request.data["reason_rent"]
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = LandRent.objects.filter(rent_year=rent_year).filter(
                reason_rent=reason_rent
            )

        # *? Empleado

        if user != 0:
            rows = (
                LandRent.objects.filter(rent_year=rent_year)
                .filter(reason_rent=reason_rent)
                .filter(user_id=user)
            )

        serializer = ListLandRentSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateLandRent(request):
    serializer = LandRentSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteLandRent(request, pk):
    landrent = LandRent.objects.get(id=pk)
    landrent.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailLandRent(request, pk):
    landrent = LandRent.objects.get(id=pk)
    serializer = ListLandRentSerializer(landrent, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateLandRent(request, pk):
    landrent = LandRent.objects.get(id=pk)
    serializer = LandRentSerializer(instance=landrent, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? CAJA CHICA
@api_view(["POST"])
def CreatePettyCash(request):
    serializer = PettyCashSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListPettyCash(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    pettycash = PettyCash.objects.filter(date__range=[fecha1, fecha2])
    serializer = ListPettyCashSerializer(pettycash, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def ListInitialCash(request):
    cash = InitialCash.objects.aggregate(Sum("cash"))
    return Response(cash)


@api_view(["POST"])
def UpdateInitialCash(request):
    cash = InitialCash.objects.get(id=1)
    serializer = InitialCashSerializer(instance=cash, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeletePettyCash(request, pk):
    pettycash = PettyCash.objects.get(id=pk)
    pettycash.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailPettyCash(request, pk):
    pettycash = PettyCash.objects.get(id=pk)
    serializer = ListPettyCashSerializer(pettycash, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdatePettyCash(request, pk):
    pettycash = PettyCash.objects.get(id=pk)
    serializer = PettyCashSerializer(instance=pettycash, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListCash(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    pettycash = Output.objects.filter(date__range=[fecha1, fecha2]).aggregate(
        Sum("total")
    )
    return Response(pettycash)
    # serializer = OutputListSerializer(pettycash, many=True)
    # return Response(serializer.data)


@api_view(["POST"])
def ListDiscount(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    discount = PettyCash.objects.filter(date__range=[fecha1, fecha2]).aggregate(
        Sum("cash")
    )
    return Response(discount)


# *? COMPRAS
@api_view(["POST"])
def Shoppings(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]

        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Shopping.objects.filter(date__range=[fecha1, fecha2])

        # *? Empleado
        if user != 0:
            rows = Shopping.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        serializer = ShoppingListSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListShopping(request):
    shopping = Shopping.objects.all()
    serializer = ShoppingListSerializer(shopping, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteShopping(request, pk):
    shopping = Shopping.objects.get(id=pk)
    shopping.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailShopping(request, pk):
    shopping = Shopping.objects.get(id=pk)
    serializer = ShoppingListSerializer(shopping, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def CreateShopping(request):
    data = request.data
    new_shopping = Shopping.objects.create(
        date=data["date"],
        provider_id=data["provider"],
        requisition_id=data["requisition"],
        department_id=data["department"],
        payment_type=data["payment_type"],
        iva=data["iva"],
        subtotal=data["subtotal"],
        total=data["total"],
        observation=data["observation"],
        applicant_id=data["applicant"],
        verify_id=data["verify"],
        authorizes_id=data["authorizes"],
        user_id=data["user"],
    )
    new_shopping.save()

    for product in data["products"]:
        new_product = ProductShopping.objects.create(
            product_name=product["product_name"],
            amount=product["amount"],
            unit=product["unit"],
            price=product["price"],
            subtotal=product["subtotal"],
            code=data["code"],
        )
        new_product.save()

        product_obj = ProductShopping.objects.filter(
            product_name=product["product_name"], code=data["code"]
        ).get(product_name=product["product_name"])
        new_shopping.products.add(product_obj)

    # serializer = RequisitionSerializer(new_shopping)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


# *? SOCIEDAD
@api_view(["POST"])
def CreateSociety(request):
    data = request.data
    new_society = Society.objects.create(
        name=data["name"], cycle=data["cycle"], year=data["year"], user_id=data["user"]
    )
    new_society.save()

    for producer in data["producers"]:
        producer_obj = Producer.objects.get(name=producer)
        new_society.producers.add(producer_obj)

    # serializer = RequisitionSerializer(new_society)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["POST"])
def UpdateSociety(request, pk):
    data = request.data
    Society.objects.filter(id=pk).update(
        name=data["name"], cycle=data["cycle"], year=data["year"]
    )
    sociedad = Society.objects.get(id=pk)
    sociedad.producers.remove()

    id = sociedad.id
    for item in data["producers"]:
        productor = Producer.objects.get(name=item)
        sociedad.producers.add(productor)

    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


@api_view(["POST"])
def Societies(request):
    if "user" in request.data:
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if user == 0:
            rows = Society.objects.all()

        # *? Empleado
        if user != 0:
            rows = Society.objects.filter(user_id=user)

        serializer = SocietySerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListSociety(request):
    society = Society.objects.all()
    serializer = SocietySerializer(society, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteSociety(request, pk):
    society = Society.objects.get(id=pk)
    society.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailSociety(request, pk):
    society = Society.objects.get(id=pk)
    serializer = SocietyListSerializer(society, many=False)
    return Response(serializer.data)


# class ReportSocietiesPDF(View):
def ReportSocietiesPDF(request, user):
    user = int(user)
    # *? administrador
    if user == 0:
        rows = Society.objects.all()
    # *? Empleado
    if user != 0:
        rows = Society.objects.filter(user_id=user)

    data = {"societies": rows}
    pdf = render_to_pdf("paginas/ReportSocieties.html", data)
    return HttpResponse(pdf, content_type="application/pdf")


class ReportSocietiesXLSX(TemplateView):
    def get(self, request, user):
        user = int(user)
        # *? administrador
        if user == 0:
            rows = Society.objects.all()
        # *? Empleado
        if user != 0:
            rows = Society.objects.filter(user_id=user)
        societies = rows
        # societies = Society.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")

        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:I2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:I3")
        ws["C5"] = "Reporte de Sociedades"
        ws.merge_cells("C5:I5")

        ws["C6"] = "ID"
        ws["D6"] = "Sociedad"
        ws["E6"] = "Ciclo"
        ws["F6"] = "Año"
        ws["G6"] = "Productores"

        cont = 7
        pos = 1

        for item in societies:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.name
            ws.cell(row=cont, column=5).value = item.cycle
            ws.cell(row=cont, column=6).value = item.year

            cont += 1

            ws.cell(row=cont, column=3).value = "PRODUCTORES"
            cont += 1
            for row in item.producers.all():
                ws.cell(row=cont, column=3).value = row.name
                cont += 1
            pos += 1

        nombre_archivo = "Reporte_Sociedades.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? SEGALMEX PARCELAS
@api_view(["GET"])
def ListSegalmexParcel(request):
    parcel = SegalmexParcel.objects.all()
    serializer = ListSegalmexParcelSerializer(parcel, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateSegalmexParcel(request):
    serializer = SegalmexParcelSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteSegalmexParcel(request, pk):
    parcel = SegalmexParcel.objects.get(id=pk)
    parcel.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailSegalmexParcel(request, pk):
    parcel = SegalmexParcel.objects.get(id=pk)
    serializer = ListSegalmexParcelSerializer(parcel, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateSegalmexParcel(request, pk):
    parcel = SegalmexParcel.objects.get(id=pk)
    serializer = SegalmexParcelSerializer(instance=parcel, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


class ReportSegalmexParcelsPDF(View):
    def get(self, request, *args, **kwargs):
        parcels = SegalmexParcel.objects.all()
        # serializer = EmployeeListSerializer(employees, many=True)
        data = {"parcels": parcels}
        pdf = render_to_pdf("paginas/ReportSegalmexParcel.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


class ReportSegalmexParcelsXLSX(TemplateView):
    def get(self, request, *args, **kwargs):
        parcels = SegalmexParcel.objects.all()
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")

        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:I2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:I3")
        ws["C5"] = "Reporte de Distribución de parcelas"
        ws.merge_cells("C5:I5")

        ws["C6"] = "ID"
        ws["D6"] = "Sociedad"
        ws["E6"] = "Ciclo PV"
        ws["F6"] = "Ciclo OI"
        ws["G6"] = "Año"
        ws["H6"] = "Comentario"

        cont = 7
        pos = 1

        for item in parcels:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.society.name
            ws.cell(row=cont, column=5).value = item.cyclepv.no_parcel
            ws.cell(row=cont, column=6).value = item.cycleoi.no_parcel
            ws.cell(row=cont, column=7).value = item.year
            ws.cell(row=cont, column=8).value = item.comment

            cont += 1
            pos += 1

        nombre_archivo = "Reporte_Sociedades.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


@api_view(["POST"])
def UploadDocProducer(request):
    serializer = FileProducerSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def DetailFileProducer(request, pk):
    documents = FileProducer.objects.filter(producer_id=pk)
    serializer = FileProducerSerializer(documents, many=True)
    return Response(serializer.data)


# *? SEGALMEX RECEPCIÓN
@api_view(["POST"])
def SearchReception(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("producer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        producer = int(request.data["producer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (producer == 0) and (user == 0):
            rows = SegalmexReception.objects.filter(
                checkin_date__range=[fecha1, fecha2]
            )

        if (producer != 0) and (user == 0):
            rows = SegalmexReception.objects.filter(producer_id=producer).filter(
                checkin_date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (producer == 0) and (user != 0):
            rows = SegalmexReception.objects.filter(user_id=user).filter(
                checkin_date__range=[fecha1, fecha2]
            )

        if (producer != 0) and (user != 0):
            rows = (
                SegalmexReception.objects.filter(user_id=user)
                .filter(producer_id=producer)
                .filter(checkin_date__range=[fecha1, fecha2])
            )

        serializer = ListSegalmexReceptionSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def ListSegalmexReception(request):
    reception = SegalmexReception.objects.all()
    serializer = ListSegalmexReceptionSerializer(reception, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateSegalmexReception(request):
    serializer = SegalmexReceptionSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteSegalmexReception(request, pk):
    reception = SegalmexReception.objects.get(id=pk)
    reception.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailSegalmexReception(request, pk):
    reception = SegalmexReception.objects.get(id=pk)
    serializer = ListSegalmexReceptionSerializer(reception, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateSegalmexReception(request, pk):
    reception = SegalmexReception.objects.get(id=pk)
    serializer = SegalmexReceptionSerializer(instance=reception, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListLocation(request):
    location = Location.objects.all()
    serializer = LocationSerializer(location, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreateVariety(request):
    serializer = VarietySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListVariety(request):
    variety = Variety.objects.all()
    serializer = VarietySerializer(variety, many=True)
    return Response(serializer.data)


class ReportReceptionsXLSX(TemplateView):
    def get(self, request, start_date, end_date, producer, user):
        fecha1 = start_date
        fecha2 = end_date
        producer = int(producer)
        user = int(user)
        # *? administrador
        if (producer == 0) and (user == 0):
            rows = SegalmexReception.objects.filter(
                checkin_date__range=[fecha1, fecha2]
            )

        if (producer != 0) and (user == 0):
            rows = SegalmexReception.objects.filter(producer_id=producer).filter(
                checkin_date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (producer == 0) and (user != 0):
            rows = SegalmexReception.objects.filter(user_id=user).filter(
                checkin_date__range=[fecha1, fecha2]
            )

        if (producer != 0) and (user != 0):
            rows = (
                SegalmexReception.objects.filter(user_id=user)
                .filter(producer_id=producer)
                .filter(checkin_date__range=[fecha1, fecha2])
            )

        receptions = rows
        wb = Workbook()
        ws = wb.active
        imag = openpyxl.drawing.image.Image("imagenes/logo_empresa_chica.png")
        ws.add_image(imag, "A1")
        ws["C2"] = "CAMPESINOS PRODUCTORES CAMPECHANOS"
        ws.merge_cells("C2:J2")
        ws["C3"] = "SPR DE RL DE CV"
        ws.merge_cells("C3:J3")
        ws["C5"] = "Reporte de SEGALMEX "
        ws.merge_cells("C5:J5")

        ws["C6"] = "ID"
        ws["D6"] = "FOLIO"
        ws["E6"] = "HORA DE ENTRADA"
        ws["F6"] = "FECHA DE ENTRADA"
        ws["G6"] = "HORA DE SALIDA"
        ws["H6"] = "FECHA DE SALIDA"
        ws["I6"] = "VARIEDAD"
        ws["J6"] = "PRODUCTOR"
        ws["K6"] = "POBLACION"
        ws["L6"] = "CHOFER"
        ws["M6"] = "PLACAS"
        ws["N6"] = "CAMION"
        ws["O6"] = "PESO(ENTRADA)BRUTO"
        ws["P6"] = "PESO TARA"
        ws["Q6"] = "PESO CAMPO"
        ws["R6"] = "% DESCUENTO"
        ws["S6"] = "PESO DE VARIEDAD"
        ws["T6"] = "PRECIO"
        ws["U6"] = "TOTAL POR BOLETA"

        cont = 7
        pos = 1
        for item in receptions:
            ws.cell(row=cont, column=3).value = pos
            ws.cell(row=cont, column=4).value = item.ticket
            ws.cell(row=cont, column=5).value = item.checkin_time
            ws.cell(row=cont, column=6).value = item.checkin_date
            ws.cell(row=cont, column=7).value = item.checkout_time
            ws.cell(row=cont, column=8).value = item.checkout_date
            ws.cell(row=cont, column=9).value = (
                item.variety.name if item.variety != None else ""
            )
            ws.cell(row=cont, column=10).value = (
                item.producer.name if item.producer != None else ""
            )
            ws.cell(row=cont, column=11).value = item.location
            ws.cell(row=cont, column=12).value = (
                item.driver.name if item.driver != None else ""
            )
            ws.cell(row=cont, column=13).value = (
                item.plate.plate_no if item.plate != None else ""
            )
            ws.cell(row=cont, column=14).value = (
                item.plate.unit if item.plate != None else ""
            )
            ws.cell(row=cont, column=15).value = item.gross_weight
            ws.cell(row=cont, column=16).value = item.tare_weight
            ws.cell(row=cont, column=17).value = item.field_weight
            ws.cell(row=cont, column=18).value = item.discount
            ws.cell(row=cont, column=19).value = item.discounted_weight
            ws.cell(row=cont, column=20).value = item.price
            ws.cell(row=cont, column=21).value = item.balance

            pos += 1
            cont += 1

        nombre_archivo = "Reporte_seglamex.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = content
        wb.save(response)
        return response


# *? PDF CLIENTES
class PDF_Boleta(View):
    def get(self, request, pk):
        boleta = SegalmexReception.objects.get(id=pk)
        data = {"boleta": boleta.id}
        pdf = render_to_pdf("paginas/Boleta.html", data)
        return HttpResponse(pdf, content_type="application/pdf")


# *? orden de pago de productores


@api_view(["GET"])
def ListPaymentOrderProducer(request):
    order = PaymentOrderProducer.objects.all()
    serializer = ListPaymentOrderProducerSerializer(order, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreatePaymentOrderProducer(request):
    serializer = PaymentOrderProducerSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeletePaymentOrderProducer(request, pk):
    order = PaymentOrderProducer.objects.get(id=pk)
    order.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailPaymentOrderProducer(request, pk):
    order = PaymentOrderProducer.objects.get(id=pk)
    serializer = ListPaymentOrderProducerSerializer(order, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdatePaymentOrderProducer(request, pk):
    order = PaymentOrderProducer.objects.get(id=pk)
    serializer = PaymentOrderProducerSerializer(instance=order, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? BUSCAR TICKET DE RECEPCIÓN
@api_view(["POST"])
def SearchTicketReception(request):
    try:
        data = request.data
        ticket = data
        reception = SegalmexReception.objects.get(ticket=ticket)
        serializer = ListSegalmexReceptionSerializer(reception, many=False)
        return Response(serializer.data)
    except SegalmexReception.DoesNotExist:
        # raise Exception("Could not find a burger")
        return Response({"message": "No se encontro coincidencia!", "status": 400})


# *? BUSCAR ORDEN DE PAGO DE PRODUCTOR
@api_view(["POST"])
def SearchPaymentOrderProducer(request):
    try:
        data = request.data
        ticket = data
        orden = PaymentOrderProducer.objects.get(id=ticket)
        serializer = ListPaymentOrderProducerSerializer(orden, many=False)
        return Response(serializer.data)
    except PaymentOrderProducer.DoesNotExist:
        # raise Exception("Could not find a burger")
        return Response({"message": "No se encontro coincidencia!", "status": 400})


# *? pago de productores


@api_view(["GET"])
def ListPaymentProducer(request):
    order = PaymentProducer.objects.all()
    serializer = ListPaymentProducerSerializer(order, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreatePaymentProducer(request):
    serializer = PaymentProducerSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeletePaymentProducer(request, pk):
    order = PaymentProducer.objects.get(id=pk)
    order.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailPaymentProducer(request, pk):
    order = PaymentProducer.objects.get(id=pk)
    serializer = ListPaymentProducerSerializer(order, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdatePaymentProducer(request, pk):
    order = PaymentProducer.objects.get(id=pk)
    serializer = PaymentProducerSerializer(instance=order, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? SALARIO DE CHOFERES


@api_view(["GET"])
def ListPlates(request):
    row = Plate.objects.all()
    serializer = PlateSerializer(row, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def CreatePlate(request):
    serializer = PlateSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListDrivers(request):
    row = Driver.objects.all()
    serializer = DriverSerializer(row, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def ListDriverSalary(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("driver" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        driver = int(request.data["driver"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (driver == 0) and (user == 0):
            rows = DriverSalary.objects.filter(date__range=[fecha1, fecha2])

        if (driver != 0) and (user == 0):
            rows = DriverSalary.objects.filter(driver_id=driver).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (driver == 0) and (user != 0):
            rows = DriverSalary.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (driver != 0) and (user != 0):
            rows = (
                DriverSalary.objects.filter(user_id=user)
                .filter(driver_id=driver)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListDriverSalarySerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateDriver(request):
    serializer = DriverSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def CreateDriverSalary(request):
    serializer = DriverSalarySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteDriverSalary(request, pk):
    driversalary = DriverSalary.objects.get(id=pk)
    driversalary.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailDriverSalary(request, pk):
    driversalary = DriverSalary.objects.get(id=pk)
    serializer = ListDriverSalarySerializer(driversalary, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateDriverSalary(request, pk):
    driversalary = DriverSalary.objects.get(id=pk)
    serializer = DriverSalarySerializer(instance=driversalary, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? SALARIO DE CARGADORES
@api_view(["POST"])
def ListChargerSalary(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("charger" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        charger = int(request.data["charger"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (charger == 0) and (user == 0):
            rows = ChargerSalary.objects.filter(date__range=[fecha1, fecha2])

        if (charger != 0) and (user == 0):
            rows = ChargerSalary.objects.filter(charger_id=charger).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (charger == 0) and (user != 0):
            rows = ChargerSalary.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (charger != 0) and (user != 0):
            rows = (
                ChargerSalary.objects.filter(user_id=user)
                .filter(charger_id=charger)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListChargerSalarySerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateChargerSalary(request):
    serializer = ChargerSalarySerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["DELETE"])
def DeleteChargerSalary(request, pk):
    chargersalary = ChargerSalary.objects.get(id=pk)
    chargersalary.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailChargerSalary(request, pk):
    chargersalary = ChargerSalary.objects.get(id=pk)
    serializer = ListChargerSalarySerializer(chargersalary, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateChargerSalary(request, pk):
    chargersalary = ChargerSalary.objects.get(id=pk)
    serializer = ChargerSalarySerializer(instance=chargersalary, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? NOMINA
@api_view(["POST"])
def CreatePayroll(request):
    data = request.data
    for item in data["rows"]:
        nomina = Payroll.objects.create(
            start_date=data["start_date"],
            end_date=data["end_date"],
            employee_id=item["employee"],
            department_id=item["department"],
            monday=item["monday"],
            tuesday=item["tuesday"],
            wednesday=item["wednesday"],
            thursday=item["thursday"],
            friday=item["friday"],
            saturday=item["saturday"],
            sunday=item["sunday"],
            worked_days=item["TotalDias"],
            hours_worked=item["TotalHoras"],
            sr=item["sr"],
            payroll=item["nomina"],
            props=item["apoyos"],
            extra_hours=item["horas_extras"],
            double_days=item["dias_dobles"],
            discounts_loan=item["descuento"],
            total_pay=item["total"],
            user_id=data["user"],
        )
        nomina.save()
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})

    """serializer = PayrollSerializer(data = request.data)
	if serializer.is_valid():
		serializer.save()
		return Response({"message":"Registro agregado satisfactoriamente!","status":200})  
	return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})"""


@api_view(["POST"])
def ListPayroll(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    nombre = Department.objects.filter(id=OuterRef("department")).values("name")
    nomina = (
        Payroll.objects.annotate(departamento=Subquery(nombre))
        .filter(start_date__range=[fecha1, fecha2])
        .annotate(TotalNomina=Sum("payroll"))
        .annotate(TotalApoyos=Sum("props"))
        .annotate(TotalHorasExtras=Sum("extra_hours"))
        .annotate(TotalDiasDobles=Sum("double_days"))
        .annotate(TotalDescuentos=Sum("discounts_loan"))
        .annotate(TotalPagos=Sum("total_pay"))
        .values(
            "departamento",
            "TotalNomina",
            "TotalApoyos",
            "TotalHorasExtras",
            "TotalDiasDobles",
            "TotalDescuentos",
            "TotalPagos",
        )
        .order_by("departamento")
    )
    # serializer = ListPayrollSerializer(nomina, many=True)
    # return Response(serializer.data)
    return Response(nomina)


@api_view(["POST"])
def CalculatePayroll(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    apoyos = (
        Props.objects.filter(employee_id=OuterRef("pk"), date__range=[fecha1, fecha2])
        .annotate(TotalProps=Sum("amount"))
        .values("TotalProps")
    )
    horas_extras = (
        ExtraHours.objects.filter(
            employee_id=OuterRef("pk"), date__range=[fecha1, fecha2]
        )
        .annotate(TotalExtrahours=Sum("hours"))
        .values("TotalExtrahours")
    )
    prestamo = (
        Loans.objects.filter(
            employee_id=OuterRef("pk"), status="PENDIENTE", date__range=[fecha1, fecha2]
        )
        .annotate(TotalAbono=Sum("payment"))
        .values("TotalAbono")
    )
    dias_dobles = (
        DoubleDays.objects.filter(
            employee_id=OuterRef("pk"), date__range=[fecha1, fecha2]
        )
        .annotate(TotalDoubledays=Count("employee_id"))
        .values("TotalDoubledays")
    )
    nomina = (
        Employee.objects.all()
        .annotate(
            nombre_completo=Concat(
                "surname",
                Value(" "),
                "second_surname",
                Value(" "),
                "name",
                output_field=models.CharField(),
            )
        )
        .annotate(TotalAbono=Subquery(prestamo))
        .annotate(TotalHorasExtras=Subquery(horas_extras))
        .annotate(TotalApoyo=Subquery(apoyos))
        .annotate(TotalDiasDobles=Subquery(dias_dobles))
        .values(
            "id",
            "nombre_completo",
            "department",
            "sr",
            "TotalAbono",
            "TotalHorasExtras",
            "TotalApoyo",
            "TotalDiasDobles",
        )
    )

    serializer = ListCalculatePayrollSerializer(nomina, many=True)
    return Response(serializer.data)

    # return Response(nomina)


# *? HORAS EXTRAS
@api_view(["POST"])
def CreateExtraHours(request):
    serializer = ExtraHoursSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListExtraHours(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    # extrahours = ExtraHours.objects.filter(date__range=[fecha1, fecha2])
    # serializer = ListExtraHoursSerializer(extrahours, many=True)
    # return Response(serializer.data)
    extrahours = (
        ExtraHours.objects.filter(date__range=[fecha1, fecha2])
        .annotate(
            nombre_completo=Concat(
                "employee__name",
                Value(" "),
                "employee__surname",
                Value(" "),
                "employee__second_surname",
                output_field=models.CharField(),
            )
        )
        .values("employee", "nombre_completo")
        .annotate(total=Count("employee"))
    )
    return Response(extrahours)


@api_view(["DELETE"])
def DeleteExtraHours(request, pk):
    extrahours = ExtraHours.objects.get(id=pk)
    extrahours.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? DIAS DOBLES
@api_view(["POST"])
def CreateDoubleDays(request):
    serializer = DoubleDaysSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListDoubleDays(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    doubledays = (
        DoubleDays.objects.filter(date__range=[fecha1, fecha2])
        .annotate(
            nombre_completo=Concat(
                "employee__name",
                Value(" "),
                "employee__surname",
                Value(" "),
                "employee__second_surname",
                output_field=models.CharField(),
            )
        )
        .values("employee", "nombre_completo")
        .annotate(total=Count("employee"))
    )
    # serializer = ListDoubleDaysSerializer(doubledays, many=True)
    # return Response(serializer.data)
    return Response(doubledays)


@api_view(["DELETE"])
def DeleteDoubleDays(request, pk):
    doubledays = DoubleDays.objects.get(id=pk)
    doubledays.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? APOYOS
@api_view(["POST"])
def CreateProps(request):
    serializer = PropsSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListProps(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    props = Props.objects.filter(date__range=[fecha1, fecha2])
    serializer = ListPropsSerializer(props, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteProps(request, pk):
    props = Props.objects.get(id=pk)
    props.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? PRESTAMOS
@api_view(["POST"])
def CreateLoans(request):
    serializer = LoansSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListLoans(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    # loans = Loans.objects.all()
    loans = (
        Loans.objects.filter(date__range=[fecha1, fecha2])
        .values("id", "date", "employee", "loan")
        .annotate(TotalPayment=Sum("payments__payment"))
        .annotate(TotalSaldo=F("loan") - F("TotalPayment"))
        .annotate(
            nombre_completo=Concat(
                "employee__name",
                Value(" "),
                "employee__surname",
                Value(" "),
                "employee__second_surname",
                output_field=models.CharField(),
            )
        )
    )
    serializer = ListLoansSerializer(loans, many=True)
    return Response(serializer.data)
    # return Response (loans)


@api_view(["DELETE"])
def DeleteLoans(request, pk):
    loans = Loans.objects.get(id=pk)
    loans.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["POST"])
def AddPaymentLoan(request):
    serializer = PaymentsSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        ActualizarStatus = Loans.objects.filter(id=request.data["loan"]).update(
            status=request.data["status"]
        )
        # ActualizarStatus.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )

    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? PRESTAMOS
@api_view(["POST"])
def CreateLoansChargers(request):
    serializer = LoansChargersSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListLoansChargers(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    user = data["user"]
    # *? administrador
    if user == 0:
        loans = LoansChargers.objects.filter(date__range=[fecha1, fecha2])
    # *? Empleado
    if user != 0:
        loans = LoansChargers.objects.filter(date__range=[fecha1, fecha2], user_id=user)

    # loans = LoansChargers.objects.filter(date__range=[fecha1, fecha2])
    serializer = ListLoansChargersSerializer(loans, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteLoansChargers(request, pk):
    loans = LoansChargers.objects.get(id=pk)
    loans.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? ABONO CARGADORES
@api_view(["POST"])
def CreatePaymentsChargers(request):
    serializer = PaymentsChargersSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListPaymentsChargers(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    user = data["user"]

    # *? administrador
    if user == 0:
        loans = PaymentsChargers.objects.filter(date__range=[fecha1, fecha2])
    # *? Empleado
    if user != 0:
        loans = PaymentsChargers.objects.filter(
            date__range=[fecha1, fecha2], user_id=user
        )

    # loans = PaymentsChargers.objects.filter(date__range=[fecha1, fecha2], user_id = user)
    serializer = ListPaymentsChargersSerializer(loans, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeletePaymentsChargers(request, pk):
    loans = PaymentsChargers.objects.get(id=pk)
    loans.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? PRESTAMOS
@api_view(["POST"])
def CreateLoansDrivers(request):
    serializer = LoansDriversSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListLoansDrivers(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    user = data["user"]
    # *? administrador
    if user == 0:
        loans = LoansDrivers.objects.filter(date__range=[fecha1, fecha2])
    # *? Empleado
    if user != 0:
        loans = LoansDrivers.objects.filter(date__range=[fecha1, fecha2], user_id=user)

    # loans = LoansDrivers.objects.filter(date__range=[fecha1, fecha2], user_id = user )
    serializer = ListLoansDriversSerializer(loans, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteLoansDrivers(request, pk):
    loans = LoansDrivers.objects.get(id=pk)
    loans.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? ABONO CHOFERES
@api_view(["POST"])
def CreatePaymentsDrivers(request):
    serializer = PaymentsDriversSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListPaymentsDrivers(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    user = data["user"]
    # *? administrador
    if user == 0:
        loans = PaymentsDrivers.objects.filter(date__range=[fecha1, fecha2])
    # *? Empleado
    if user != 0:
        loans = PaymentsDrivers.objects.filter(
            date__range=[fecha1, fecha2], user_id=user
        )

    # loans = PaymentsDrivers.objects.filter(date__range=[fecha1, fecha2], user_id = user)
    serializer = ListPaymentsDriversSerializer(loans, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeletePaymentsDrivers(request, pk):
    loans = PaymentsDrivers.objects.get(id=pk)
    loans.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? CARTA PORTE
@api_view(["POST"])
def BillOfLadings(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = BillOfLading.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = BillOfLading.objects.filter(customer_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = BillOfLading.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                BillOfLading.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListBillOfLadingSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateBillOfLading(request):
    serializer = BillOfLadingSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListBillOfLading(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    bill = BillOfLading.objects.filter(date__range=[fecha1, fecha2])
    serializer = ListBillOfLadingSerializer(bill, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteBillOfLading(request, pk):
    bill = BillOfLading.objects.get(id=pk)
    bill.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def DetailBillOfLading(request, pk):
    row = BillOfLading.objects.get(id=pk)
    serializer = ListBillOfLadingSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateBillOfLading(request, pk):
    row = BillOfLading.objects.get(id=pk)
    serializer = BillOfLadingSerializer(instance=row, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? PROGRAMACIÓN DE PAGOS
@api_view(["POST"])
def PaymentsSchedule(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = PaymentSchedule.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = PaymentSchedule.objects.filter(customer_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = PaymentSchedule.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                PaymentSchedule.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListPaymentScheduleSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreatePaymentSchedule(request):
    data = request.data
    new_payment = PaymentSchedule.objects.create(
        date=data["date"],
        company_id=data["company"],
        customer_id=data["customer"],
        bank_account_id=data["bank_account"],
        description=data["description"],
        total=data["total"],
        code=data["code"],
        user_id=data["user"],
    )
    new_payment.save()

    for product in data["concepts"]:
        new_product = ConceptPayment.objects.create(
            concept=product["concept"], amount=product["amount"], code=data["code"]
        )
        new_product.save()

        product_obj = ConceptPayment.objects.filter(
            concept=product["concept"], amount=product["amount"], code=data["code"]
        ).get(code=data["code"])
        new_payment.concepts.add(product_obj)

    # serializer = RequisitionSerializer(new_creditnote)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["POST"])
def ListPaymentSchedule(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    paymentschedule = PaymentSchedule.objects.filter(date__range=[fecha1, fecha2])
    serializer = ListPaymentScheduleSerializer(paymentschedule, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeletePaymentSchedule(request, pk):
    payment = PaymentSchedule.objects.get(id=pk)
    payment.delete()
    return Response("Registro eliminado satisfactoriamente!")


# *? CUENTAS DE BANCOS DE CLIENTES


@api_view(["POST"])
def CreateBankAccountsCustomer(request):
    serializer = BankAccountsCustomerSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListBankAccountsCustomer(request, pk):
    list = BankAccountsCustomer.objects.filter(customer_id=pk)
    serializer = ListBankAccountsCustomerSerializer(list, many=True)
    return Response(serializer.data)


# *? CUENTAS DE BANCOS DE EMPLEADOS


@api_view(["POST"])
def CreateBankAccountsEmployee(request):
    serializer = BankAccountsEmployeeSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListBankAccountsEmployee(request, pk):
    list = BankAccountsEmployee.objects.filter(employee_id=pk)
    serializer = ListBankAccountsEmployeeSerializer(list, many=True)
    return Response(serializer.data)


# *? CUENTAS DE BANCOS DE PROVEEDORES


@api_view(["POST"])
def CreateBankAccountsProvider(request):
    serializer = BankAccountsProviderSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["GET"])
def ListBankAccountsProvider(request, pk):
    list = BankAccountsProvider.objects.filter(provider_id=pk)
    serializer = ListBankAccountsProviderSerializer(list, many=True)
    return Response(serializer.data)


# *? FACTURAS
@api_view(["POST"])
def ListBilledIncome(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = BilledIncome.objects.filter(date__range=[fecha1, fecha2]).order_by(
                "invoice"
            )

        if (customer != 0) and (user == 0):
            rows = (
                BilledIncome.objects.filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
                .order_by("invoice")
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = (
                BilledIncome.objects.filter(user_id=user)
                .filter(date__range=[fecha1, fecha2])
                .order_by("invoice")
            )

        if (customer != 0) and (user != 0):
            rows = (
                BilledIncome.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            ).order_by("invoice")

        serializer = ListBilledIncomeSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["POST"])
def CreateBilledIncome(request):
    data = request.data
    nueva_factura = BilledIncome.objects.create(
        date=data["date"],
        customer_id=data["customer"],
        due_date=data["due_date"],
        invoice=data["invoice"],
        net_weight=data["net_weight"],
        total=data["total"],
        user_id=data["user"],
        status=data["status"],
    )
    nueva_factura.save()

    for product in data["products"]:
        new_product = ProductBI.objects.create(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            net_weight=product["net_weight"],
            price=product["price"],
            subtotal=product["subtotal"],
        )
        new_product.save()
        product_obj = ProductBI.objects.get(id=new_product.id)
        # product_obj = ProductPO.objects.filter(amount=product["amount"], unit=product["unit"], presentation_id=product["presentation"], product_id=product["product"], price=product["price"], subtotal = product["subtotal"], code=data["code"]).get(code=data["code"])
        nueva_factura.products.add(product_obj)

    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})


@api_view(["GET"])
def ListBills(request):
    list = BilledIncome.objects.all()
    serializer = ListBilledIncomeSerializer(list, many=True)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteBilledIncome(request, pk):
    row = BilledIncome.objects.filter(id=pk)
    row.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


@api_view(["GET"])
def DetailBilledIncome(request, pk):
    row = BilledIncome.objects.get(id=pk)
    serializer = ListBilledIncomeSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateBilledIncome(request, pk):
    data = request.data
    BilledIncome.objects.filter(id=pk).update(
        date=data["date"],
        customer_id=data["customer"],
        due_date=data["due_date"],
        invoice=data["invoice"],
        net_weight=data["net_weight"],
        total=data["total"],
        user_id=data["user"],
        status=data["status"],
    )
    for product in data["rows"]:
        new_product = ProductBI.objects.filter(id=product["id"]).update(
            amount=product["amount"],
            unit=product["unit"],
            presentation_id=product["presentation"],
            product_id=product["product"],
            net_weight=product["net_weight"],
            price=product["price"],
            subtotal=product["subtotal"],
        )
    # new_product.save()
    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


# *? INGRESOS
@api_view(["POST"])
def CreateIncome(request):
    data = request.data
    new_income = Income.objects.create(
        date=data["date"],
        customer_id=data["customer"],
        bank_account_id=data["bank_account"],
        payment_plugin_id=data["payment_plugin"],
        amount=data["amount"],
        user_id=data["user"],
    )
    new_income.save()

    id = new_income.id

    for factura in data["invoices"]:
        ##myfac = int(factura)
        bill = BilledIncome.objects.get(invoice=factura)
        # id_factura = 5
        new_invoices = IncomeInvoices.objects.create(
            invoice_id=bill.id, income=id, user_id=data["user"]
        )
        new_invoices.save()
        # factura_obj = IncomeInvoices.objects.filter(income=id).get(invoice=factura)
        factura_obj = IncomeInvoices.objects.get(id=new_invoices.id)
        new_income.invoices.add(factura_obj)

    for notas in data["creditnotes"]:
        note = CreditNote.objects.get(no_creditnote=notas)
        new_notes = IncomeCreditNotes.objects.create(
            creditnote_id=note.id, income=id, user_id=data["user"]
        )
        new_notes.save()
        # notas_obj = IncomeCreditNotes.objects.filter(income=id).get(creditnote=notas)
        notas_obj = IncomeCreditNotes.objects.get(id=new_notes.id)
        new_income.creditnotes.add(notas_obj)

    # serializer = RequisitionSerializer(new_creditnote)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["POST"])
def ListFactura(request):
    bill = BilledIncome.objects.get(invoice=request.data["factura"])
    # serializer = ListBilledIncomeSerializer(bill, many=False)
    # return Response(serializer.data)
    return Response({"consulta": bill.id})


@api_view(["POST"])
def ListIncome(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = Income.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = Income.objects.filter(customer_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = Income.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                Income.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListIncomeSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def DetailIncome(request, pk):
    row = Income.objects.get(id=pk)
    serializer = ListIncomeSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeleteIncome(request, pk):
    row = Income.objects.get(id=pk)
    items = row.invoices.all()
    for item in items:
        producto = IncomeInvoices.objects.get(id=item.id)
        producto.delete()

    items = row.creditnotes.all()
    for item in items:
        producto = IncomeCreditNotes.objects.get(id=item.id)
        producto.delete()

    row.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["POST"])
def UpdateIncome(request, pk):
    data = request.data
    #
    Income.objects.filter(id=pk).update(
        date=data["date"],
        customer_id=data["customer"],
        bank_account_id=data["bank_account"],
        payment_plugin=data["payment_plugin"],
        amount=data["amount"],
    )
    row = Income.objects.get(id=pk)
    items = row.invoices.all()
    for item in items:
        producto = IncomeInvoices.objects.get(id=item.id)
        producto.delete()
    items = row.creditnotes.all()
    for item in items:
        producto = IncomeCreditNotes.objects.get(id=item.id)
        producto.delete()

    id = row.id
    for factura in data["invoices"]:
        ##myfac = int(factura)
        bill = BilledIncome.objects.get(invoice=factura)
        # id_factura = 5
        new_invoices = IncomeInvoices.objects.create(
            invoice_id=bill.id, income=id, user_id=data["user"]
        )
        new_invoices.save()
        # factura_obj = IncomeInvoices.objects.filter(income=id).get(invoice=factura)
        factura_obj = IncomeInvoices.objects.get(id=new_invoices.id)
        row.invoices.add(factura_obj)

    for notas in data["creditnotes"]:
        note = CreditNote.objects.get(no_creditnote=notas)
        new_notes = IncomeCreditNotes.objects.create(
            creditnote_id=note.id, income=id, user_id=data["user"]
        )
        new_notes.save()
        # notas_obj = IncomeCreditNotes.objects.filter(income=id).get(creditnote=notas)
        notas_obj = IncomeCreditNotes.objects.get(id=new_notes.id)
        row.creditnotes.add(notas_obj)

    return Response(
        {"message": "Registro actualizado satisfactoriamente!", "status": 200}
    )


# *? VENTAS


@api_view(["POST"])
def CreateSale(request):
    serializer = SaleSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListSales(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = Sale.objects.filter(date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = Sale.objects.filter(customer_id=customer).filter(
                date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = Sale.objects.filter(user_id=user).filter(
                date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                Sale.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(date__range=[fecha1, fecha2])
            )

        serializer = ListSalesSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["DELETE"])
def DeleteSale(request, pk):
    row = Sale.objects.filter(id=pk)
    row.delete()
    return Response(
        {"message": "Registro eliminado satisfactoriamente!", "status": 200}
    )


@api_view(["GET"])
def DetailSale(request, pk):
    row = Sale.objects.get(id=pk)
    serializer = SaleSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["POST"])
def UpdateSale(request, pk):
    row = Sale.objects.get(id=pk)
    serializer = SaleSerializer(instance=row, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro actualizado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó la actualización!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


# *? CONTROL DE DEPOSITOS


@api_view(["POST"])
def CreateDepositControl(request):
    serializer = DepositControlSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListDepositControl(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    account = data["account"]
    list = DepositControl.objects.filter(date__range=[fecha1, fecha2]).filter(
        account_id=account
    )
    serializer = ListDepositControlSerializer(list, many=True)
    return Response(serializer.data)


# *? CONCILIACION


@api_view(["POST"])
def CreateConciliation(request):
    serializer = ConciliationSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(
            {"message": "Registro agregado satisfactoriamente!", "status": 200}
        )
    return Response(
        {
            "message": "No se realizó el Registro!",
            "errors": serializer.errors,
            "status": 400,
        }
    )


@api_view(["POST"])
def ListConciliation(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]
    account = data["account"]
    list = Conciliation.objects.filter(date__range=[fecha1, fecha2]).filter(
        account_id=account
    )
    serializer = ListConciliationSerializer(list, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def ListConciliationForMonth(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    nombre = BankAccount.objects.filter(id=OuterRef("account")).values("company")
    # saldo_inicial = Conciliation.objects.filter(date__range=[fecha1, fecha2]).values('countable_balance').order_by('id').first()
    list = (
        Conciliation.objects.filter(date__range=[fecha1, fecha2])
        .annotate(RazonSocial=Subquery(nombre))
        .annotate(Ingresos=Sum("deposit"))
        .annotate(Egresos=Sum("charge"))
        .values("account", "RazonSocial", "Ingresos", "Egresos")
        .order_by("account")
    )

    # list = Conciliation.objects.filter(date__range=[fecha1, fecha2]).annotate(SaldoInicial = Min('id')).values('SaldoInicial')

    return Response(list)


@api_view(["DELETE"])
def DeleteConciliation(request, pk):
    row = Conciliation.objects.get(id=pk)
    row.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["GET"])
def PreviousLedgerBalance(request, pk):
    data = request.data
    account = pk
    try:
        ultimo_saldo = (
            Conciliation.objects.filter(account_id=account)
            .latest("id")
            .countable_balance
        )
        return Response(ultimo_saldo)
    except Conciliation.DoesNotExist:
        return Response(0)
    # return Response(account)


@api_view(["GET"])
def PreviousAccountBalance(request, pk):
    data = request.data
    account = pk
    try:
        ultimo_saldo = (
            Conciliation.objects.filter(account_id=account).latest("id").account_balance
        )
        return Response(ultimo_saldo)
    except Conciliation.DoesNotExist:
        return Response(0)
    # return Response(account)


@api_view(["POST"])
def ListTotalDepositControl(request):
    data = request.data
    fecha1 = data["start_date"]
    fecha2 = data["end_date"]

    list = (
        DepositControl.objects.filter(date__range=[fecha1, fecha2])
        .values("account__tax_regime", "account__company", "account__deposit_limit")
        .order_by("account__tax_regime", "account", "account__deposit_limit")
        .annotate(TotalAmount=Sum("amount"))
    )
    # serializer = ListTotalDepositControlSerializer(list, many=True)
    # return Response(serializer.data)
    return Response(list)


# *? COMPLEMENTOS DE PAGO


@api_view(["POST"])
def CreatePaidPlugins(request):
    data = request.data
    # archivo = request.FILES.get("document")
    # facturas = data["bills"]
    new_paid = PaidPlugins.objects.create(
        invoice=data["invoice"],
        payment_date=data["payment_date"],
        date_issue=data["date_issue"],
        # company_id = data['company'],
        customer_id=data["customer"],
        account_id=data["account"],
        amount=data["amount"],
        # code=data["code"],
        user_id=data["user"],
        document=data["document"],
    )

    new_paid.save()
    bills = request.data.get("bills", [])
    if bills:
        bills = json.loads(bills)

        # new_paid.id

        # complemento = PaidPlugins.objects.get(id=new_paid.id)
        # complemento.document = data["document"]
        # complemento.save()

        for row in bills:
            bill = BilledIncome.objects.get(invoice=row)
            new_invoice = BillsPaidPlugins.objects.create(
                invoice_id=bill.id, customer_id=data["customer"]
            )
            new_invoice.save()

            invoice_obj = BillsPaidPlugins.objects.get(id=new_invoice.id)
            new_paid.bills.add(invoice_obj)

    # serializer = RequisitionSerializer(new_creditnote)
    # return Response(serializer.data)
    return Response({"message": "Registro agregado satisfactoriamente!", "status": 200})
    # return Response({"message":"No se realizó el Registro!","errors":serializer.errors,"status":400})


@api_view(["GET"])
def ListPlugins(request):
    rows = PaidPlugins.objects.all()
    serializer = ListPaidPluginsSerializer(rows, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def ListPaidPlugins(request):
    if (
        ("start_date" in request.data)
        and ("end_date" in request.data)
        and ("customer" in request.data)
        and ("user" in request.data)
    ):
        fecha1 = request.data["start_date"]
        fecha2 = request.data["end_date"]
        customer = int(request.data["customer"])
        user = int(request.data["user"])
        # return Response({"department->":department,"user->":user})
        # *? administrador
        if (customer == 0) and (user == 0):
            rows = PaidPlugins.objects.filter(payment_date__range=[fecha1, fecha2])

        if (customer != 0) and (user == 0):
            rows = PaidPlugins.objects.filter(customer_id=customer).filter(
                payment_date__range=[fecha1, fecha2]
            )

        # *? Empleado
        if (customer == 0) and (user != 0):
            rows = PaidPlugins.objects.filter(user_id=user).filter(
                payment_date__range=[fecha1, fecha2]
            )

        if (customer != 0) and (user != 0):
            rows = (
                PaidPlugins.objects.filter(user_id=user)
                .filter(customer_id=customer)
                .filter(payment_date__range=[fecha1, fecha2])
            )

        serializer = ListPaidPluginsSerializer(rows, many=True)
        return Response(serializer.data)

    else:
        return Response({"mensaje": "sin variables"})


@api_view(["GET"])
def DetailPaidPlugins(request, pk):
    row = PaidPlugins.objects.get(id=pk)
    serializer = ListPaidPluginsSerializer(row, many=False)
    return Response(serializer.data)


@api_view(["DELETE"])
def DeletePaidPlugins(request, pk):
    row = PaidPlugins.objects.get(id=pk)
    items = row.bills.all()
    for item in items:
        producto = BillsPaidPlugins.objects.get(id=item.id)
        producto.delete()
    row.delete()
    return Response("Registro eliminado satisfactoriamente!")


@api_view(["POST"])
def ConvertToLetter(request):
    if "balance" in request.data:
        cantidad = float(request.data["balance"])
        # cantidad = 15000.50
        # cantidad_entero = int(cantidad)
        # cantidad_decimal = int(round((cantidad - cantidad_entero) * 100))
        # cantidad_texto = f"{num2words(cantidad_entero, lang='es')} punto {num2words(cantidad_decimal, lang='es')}"
        # return Response(cantidad_decimal)
        # Redondeamos la cantidad a dos decimales
        cantidad = round(cantidad, 2)

        # Convertimos la parte entera a texto
        parte_entera = num2words(int(cantidad), lang="es").capitalize()

        # Obtenemos los decimales como una cadena de texto
        decimales = str(cantidad).split(".")[1]

        # Convertimos los decimales a texto
        parte_decimal = num2words(int(decimales), lang="es").capitalize()

        # Creamos la representación en texto de la cantidad
        if decimales == "00":
            cantidad_en_texto = f"{parte_entera} "
        else:
            cantidad_en_texto = f"{parte_entera} punto {parte_decimal}"

        # return cantidad_en_texto

        # datos = {"texto": cantidad_en_texto}

        # Convertir el diccionario a formato JSON
        # json_texto = json.dumps(datos)

        # print(json_texto)
        return Response({"message": cantidad_en_texto, "status": 200})
        # return Response(json_texto)
